"""Load the operator's planned-settlements workbook into mda_planned_settlements.

One-shot loader. Reads an XLSX with columns (in the shape produced by the
CommCare planning workbook):

    form village_location admin2                             -> lga
    form village_location ward_name admin3                   -> ward_name
    form village_location ward_name admin3_code              -> admin3_code
    form village_location settlement_name admin5             -> settlement_name
    form village_location settlement_name admin5_code        -> admin5_code

Idempotent:
  * Ensures the ``mda_planned_settlements`` table exists (raw DDL matching
    the SQLAlchemy model). Safe to run against a DB where the model has
    already been created via Base.metadata.create_all.
  * Uses ``ON CONFLICT (project_id, admin5_code) DO UPDATE`` so re-running
    for the same project overwrites without duplicating rows.

Usage
-----
    # Kano R3 (project_id 4) against prod RDS via the SSM tunnel
    python scripts/load_planned_settlements.py \\
        --project-id 4 \\
        --xlsx planned_lga_ward_settlement_list.xlsx

Environment
-----------
Reads DB creds from DATABASE_URL_SYNC in .env (same as the CommCare sync
worker). Bring the tunnel up with ``bash scripts/dev-aws.sh up`` first.
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from typing import Iterable, Tuple

# Third-party
import openpyxl
import psycopg2
import psycopg2.extras


DDL = """
CREATE TABLE IF NOT EXISTS mda_planned_settlements (
    id              SERIAL PRIMARY KEY,
    project_id      INTEGER NOT NULL REFERENCES geo_projects(id),
    lga             TEXT,
    ward_name       TEXT,
    admin3_code     TEXT,
    settlement_name TEXT,
    admin5_code     TEXT NOT NULL,
    uploaded_at     TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
    CONSTRAINT uq_planned_settlement_project_code UNIQUE (project_id, admin5_code)
);
CREATE INDEX IF NOT EXISTS ix_mda_planned_settlements_project_id ON mda_planned_settlements (project_id);
CREATE INDEX IF NOT EXISTS ix_mda_planned_settlements_lga        ON mda_planned_settlements (lga);
CREATE INDEX IF NOT EXISTS ix_mda_planned_settlements_admin3     ON mda_planned_settlements (admin3_code);
CREATE INDEX IF NOT EXISTS ix_mda_planned_settlements_admin5     ON mda_planned_settlements (admin5_code);
"""


COLUMN_MAP = {
    "form village_location admin2":                       "lga",
    "form village_location ward_name admin3":             "ward_name",
    "form village_location ward_name admin3_code":        "admin3_code",
    "form village_location settlement_name admin5":       "settlement_name",
    "form village_location settlement_name admin5_code":  "admin5_code",
}


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _read_workbook(path: Path) -> Iterable[Tuple[str, str, str, str, str]]:
    """Yield (lga, ward_name, admin3_code, settlement_name, admin5_code) per row.

    Reads the first sheet, expects the CommCare-workbook headers listed at
    the top of this file. Rows missing ``admin5_code`` are skipped: without
    the code we can't match against mda_households, so the row is unusable.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    idx_map = {}
    skipped_no_code = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [(c or "").strip() for c in row]
            for j, h in enumerate(header):
                if h in COLUMN_MAP:
                    idx_map[COLUMN_MAP[h]] = j
            missing = [k for k in COLUMN_MAP.values() if k not in idx_map]
            if missing:
                raise SystemExit(
                    f"Workbook is missing required columns for {missing}. "
                    f"Actual headers: {header}"
                )
            continue
        lga            = _clean(row[idx_map["lga"]])
        ward_name      = _clean(row[idx_map["ward_name"]])
        admin3_code    = _clean(row[idx_map["admin3_code"]])
        settlement     = _clean(row[idx_map["settlement_name"]])
        admin5_code    = _clean(row[idx_map["admin5_code"]])
        if not admin5_code:
            skipped_no_code += 1
            continue
        yield lga, ward_name, admin3_code, settlement, admin5_code
    if skipped_no_code:
        print(f"  skipped {skipped_no_code:,} rows with no admin5_code (unusable for matching)")


def _get_db_url() -> str:
    """Read DATABASE_URL_SYNC from .env if present; otherwise the env var."""
    url = os.environ.get("DATABASE_URL_SYNC")
    if url:
        return url
    env_path = Path(".env")
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if line.startswith("DATABASE_URL_SYNC="):
                return line.split("=", 1)[1].strip()
    raise SystemExit(
        "DATABASE_URL_SYNC not set. Run `bash scripts/dev-aws.sh up` to write "
        ".env with the tunnel-scoped connection string, then re-run."
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument("--project-id", type=int, required=True,
                        help="geo_projects.id — Kano R3 is 4")
    parser.add_argument("--xlsx", type=Path, required=True,
                        help="Path to the planned-settlements workbook")
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"XLSX not found: {args.xlsx}")

    db_url = _get_db_url()
    # psycopg2 doesn't accept SQLAlchemy's '+asyncpg' / '+psycopg' driver
    # prefixes. Strip them so the same DATABASE_URL_SYNC line works whether
    # it was written for asyncpg (dev API) or plain psycopg2 (this script).
    if db_url.startswith("postgresql+"):
        db_url = "postgresql://" + db_url.split("://", 1)[1]
    # scripts/dev-aws.sh writes the URL for the docker container, using
    # ``host.docker.internal`` so the containerised API can reach the tunnel
    # on the host. This script runs on the host directly - Windows resolves
    # ``host.docker.internal`` to the Docker gateway IP, which times out.
    # Rewrite to localhost so the host-side psycopg2 connect works.
    db_url = db_url.replace("@host.docker.internal:", "@localhost:")

    print(f"Reading workbook: {args.xlsx}")
    rows = list(_read_workbook(args.xlsx))
    print(f"  parsed rows with admin5_code: {len(rows):,}")

    if not rows:
        raise SystemExit("Nothing to load.")

    print(f"Connecting to DB (target project_id={args.project_id})...")
    conn = psycopg2.connect(db_url)
    conn.autocommit = False
    cur = conn.cursor()

    # Ensure the target project exists — refuse to insert with a bad FK
    cur.execute("SELECT name FROM geo_projects WHERE id = %s", (args.project_id,))
    proj = cur.fetchone()
    if not proj:
        raise SystemExit(f"No geo_projects row with id={args.project_id}")
    print(f"  target project: '{proj[0]}'")

    print("Ensuring table + indexes exist...")
    cur.execute(DDL)

    # The loader connects as the table owner (app_dev via the dev tunnel), so a
    # freshly-created table is NOT readable by the live app role (app_prod) until
    # explicitly granted. Without this, the geo endpoints + planned-vs-reached
    # endpoints 500 in production ("permission denied for table
    # mda_planned_settlements"). Grant read to app_prod here so it can never
    # recur on a re-create or a new environment. Guarded so it's a no-op where
    # the app_prod role doesn't exist (e.g. a local-only dev DB).
    print("Granting SELECT to app_prod (if the role exists)...")
    cur.execute("""
        DO $$
        BEGIN
          IF EXISTS (SELECT 1 FROM pg_roles WHERE rolname = 'app_prod') THEN
            GRANT SELECT ON mda_planned_settlements TO app_prod;
          END IF;
        END $$;
    """)

    print(f"Upserting {len(rows):,} rows...")
    payload = [(args.project_id, lga, ward, a3, sett, a5)
               for (lga, ward, a3, sett, a5) in rows]
    psycopg2.extras.execute_values(
        cur,
        """
        INSERT INTO mda_planned_settlements
            (project_id, lga, ward_name, admin3_code, settlement_name, admin5_code)
        VALUES %s
        ON CONFLICT (project_id, admin5_code) DO UPDATE SET
            lga             = EXCLUDED.lga,
            ward_name       = EXCLUDED.ward_name,
            admin3_code     = EXCLUDED.admin3_code,
            settlement_name = EXCLUDED.settlement_name,
            uploaded_at     = CURRENT_TIMESTAMP
        """,
        payload,
        page_size=1000,
    )

    # Sanity — count what actually landed
    cur.execute(
        "SELECT COUNT(*) FROM mda_planned_settlements WHERE project_id = %s",
        (args.project_id,),
    )
    n_planned = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(DISTINCT p.admin5_code) AS visited
        FROM mda_planned_settlements p
        WHERE p.project_id = %s
          AND EXISTS (
            SELECT 1 FROM mda_households h
            WHERE h.project_id = p.project_id AND h.admin5_code = p.admin5_code
          )
    """, (args.project_id,))
    n_visited = cur.fetchone()[0]

    n_not_visited = n_planned - n_visited
    conn.commit()
    cur.close()
    conn.close()

    pct = (100.0 * n_visited / n_planned) if n_planned else 0.0
    print()
    print(f"Total planned settlements now in DB: {n_planned:,}")
    print(f"  Visited     (>=1 form with matching admin5_code): {n_visited:,} ({pct:.1f}%)")
    print(f"  Not visited (0 matching forms)                 : {n_not_visited:,} ({100.0 - pct:.1f}%)")


if __name__ == "__main__":
    sys.exit(main())
