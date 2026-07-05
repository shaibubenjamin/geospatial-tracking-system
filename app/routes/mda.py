"""
app/routes/mda.py — MDA Data Quality Check System
Upload endpoint + QC query endpoints for Mass Drug Administration data.
"""
import io
import logging
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional

import psycopg2
import psycopg2.extras
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from app.database import get_db
from app.models import User
from app.routes.auth import get_current_user, get_current_user_optional, require_admin, require_superadmin, allowed_states_of, allowed_lgas_of
from app.config import DATABASE_URL_SYNC

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/mda", tags=["mda"])


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _str(val) -> Optional[str]:
    """Convert cell value to stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s not in ("", "---", "None") else None


def _int_safe(val) -> Optional[int]:
    try:
        return int(float(str(val)))
    except (TypeError, ValueError):
        return None


def _float_safe(val) -> Optional[float]:
    try:
        return float(str(val))
    except (TypeError, ValueError):
        return None


def _map_columns(header_row: List, schema: Dict[str, tuple]) -> Dict[str, Optional[int]]:
    """Build a {target_field → column_index} map from a workbook header row.

    Designed for upload endpoints (MLOS, baseline, MDA) that receive
    spreadsheets whose column names vary between data partners ("State"
    vs "state_name", "Ward" vs "wardname", "Lat" vs "latitude" …). For
    each target field the caller passes a tuple of candidate names; the
    first column whose header (case-insensitively) equals one of the
    candidates wins. A substring match falls back next (so "wardname"
    matches "ward_name" or "wardName").

    Returns a dict mapping every target field to either a column index
    or ``None`` if no header matched any candidate. Callers can detect
    missing required columns by checking for ``None``.
    """
    norm = [(str(h or "").strip().lower(), i) for i, h in enumerate(header_row)]
    out: Dict[str, Optional[int]] = {}
    for field, candidates in schema.items():
        idx: Optional[int] = None
        for cand in candidates:
            cand_l = cand.lower()
            # exact case-insensitive match
            for hdr_l, i in norm:
                if cand_l == hdr_l:
                    idx = i
                    break
            if idx is not None:
                break
            # substring fallback
            for hdr_l, i in norm:
                if cand_l in hdr_l:
                    idx = i
                    break
            if idx is not None:
                break
        out[field] = idx
    return out


def _parse_gps(raw: Optional[str]):
    """
    Parse '12.7209383 5.0116833 343.2 4.58' → (lat, lon, alt, accuracy).
    Returns (None, None, None, None) on any error.
    """
    if not raw:
        return None, None, None, None
    parts = str(raw).strip().split()
    try:
        lat = float(parts[0])
        lon = float(parts[1])
        alt = float(parts[2]) if len(parts) > 2 else None
        acc = float(parts[3]) if len(parts) > 3 else None
        return lat, lon, alt, acc
    except (IndexError, ValueError):
        return None, None, None, None


def _parse_dt(val) -> Optional[datetime]:
    """Parse a datetime cell to UTC-aware datetime or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        if val.tzinfo is None:
            return val.replace(tzinfo=timezone.utc)
        return val.astimezone(timezone.utc)
    s = str(val).strip()
    if not s or s in ("---", "None"):
        return None
    # Try common ISO formats
    for fmt in (
        "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            dt = datetime.strptime(s, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except ValueError:
            continue
    return None


def _parse_date(val) -> Optional[str]:
    """Return ISO date string or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    if not s or s in ("---", "None"):
        return None
    # strip time portion if present
    return s.split("T")[0].split(" ")[0] if len(s) >= 10 else None


def _get_sync_conn():
    """Open a synchronous psycopg2 connection using DATABASE_URL_SYNC."""
    return psycopg2.connect(DATABASE_URL_SYNC)


async def resolve_pid(
    project_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    user: "Optional[User]" = Depends(get_current_user_optional),
) -> int:
    """Resolve which project's data to query, enforcing per-user STATE access.

    - Explicit ?project_id=N → that project, but a state-scoped (non-superadmin)
      user gets 403 if it belongs to a state they're not assigned to.
    - Otherwise → the active project within the user's state(s); falls back to
      the global active project for superadmins and the public dashboard.
    - Last resort → the lowest-id project (fresh DB).

    This is the single choke-point every project-scoped endpoint depends on, so
    state access is enforced identically for the web platform and the app.
    """
    scoped = user is not None and not user.is_superadmin and not user.is_admin
    allowed = allowed_states_of(user) if scoped else None  # a set for scoped users

    if user is None:
        # Anonymous (public dashboard) → only projects opted in as public.
        if project_id is not None:
            is_pub = (await db.execute(
                text("SELECT COALESCE(is_public, FALSE) FROM geo_projects WHERE id = :id"),
                {"id": project_id},
            )).scalar()
            if not is_pub:
                raise HTTPException(status_code=403, detail="This campaign isn't public.")
            return project_id
        row = (await db.execute(text(
            "SELECT id FROM geo_projects WHERE COALESCE(is_public, FALSE) = TRUE "
            "ORDER BY is_active DESC, round_number DESC NULLS LAST, id LIMIT 1"
        ))).fetchone()
        if row:
            return row[0]
        # No public project configured — fall through to the global default.

    if project_id is not None:
        if scoped:
            state = (await db.execute(
                text("SELECT state_name FROM geo_projects WHERE id = :id"),
                {"id": project_id},
            )).scalar()
            if (state or "").strip().lower() not in (allowed or set()):
                raise HTTPException(status_code=403, detail="You don't have access to this campaign.")
        return project_id

    if scoped:
        if allowed:
            row = (await db.execute(
                text("SELECT id FROM geo_projects WHERE lower(state_name) = ANY(:st) "
                     "ORDER BY is_active DESC, round_number DESC NULLS LAST, id LIMIT 1"),
                {"st": list(allowed)},
            )).fetchone()
            if row:
                return row[0]
        # A scoped user with no project in their state(s) sees NOTHING — never
        # fall through to another state's data.
        raise HTTPException(status_code=403, detail="No campaign is available for your assigned state(s).")

    # Superadmin / anonymous (public) / unrestricted → global active project.
    res = await db.execute(text("SELECT id FROM geo_projects WHERE is_active = TRUE ORDER BY id LIMIT 1"))
    row = res.fetchone()
    if row:
        return row[0]
    res = await db.execute(text("SELECT id FROM geo_projects ORDER BY id LIMIT 1"))
    row = res.fetchone()
    return row[0] if row else 1


def _lga_and(lgas, col: str, params: dict, key: str = "lgascope") -> str:
    """SQL fragment restricting an LGA-name column to an LGA-scoped user's set.

    - ``lgas`` is ``None`` (no LGA restriction — admins, public, or a state-only
      user) → returns "" so the query is unchanged.
    - a non-empty set → ``AND lower(trim(col)) = ANY(:lgascope)``.
    - an empty set (an LGA account whose list is blank/malformed) → ``AND FALSE``
      so it fails CLOSED (no rows) rather than leaking every LGA.

    The same ``key`` is reused across all sub-queries of one statement (the bound
    value is identical), so callers can drop this into several predicates.
    """
    if lgas is None:
        return ""
    if not lgas:
        return " AND FALSE"
    params[key] = list(lgas)
    return f" AND lower(trim({col})) = ANY(:{key})"


def _as_date(s: "Optional[str]"):
    """Parse a YYYY-MM-DD filter string into a datetime.date.

    asyncpg binds a date-typed query parameter as a datetime.date, NOT a str, so
    passing the raw 'YYYY-MM-DD' string 500s every date-filtered query with
    "'str' object has no attribute 'toordinal'". Callers convert here before
    putting the value in params. Returns None for empty/malformed input (the
    filter then matches nothing instead of crashing the request)."""
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _scoped_where(pid: int, filters: list, params: dict, alias: str = "",
                  date_filter: bool = True, date_col: str = "received_on",
                  lgas=None, lga_col: str = "lga") -> str:
    """Build a WHERE clause that always pins the query to a single project.

    When ``date_filter`` is True (default) we also bound the query at the
    project's ``campaign_start_date`` if one is set on geo_projects. This
    lets the campaign team mark the official Day-1 (e.g. R5 starts 19 May)
    and have every Days-Active / daily-trends / coverage-pace metric hide
    pre-campaign test submissions without touching the raw data.

    Pass ``alias`` when the table is aliased (e.g. ``"h"`` for ``mda_households h``).
    Pass ``date_filter=False`` for queries that don't reference received_on.
    Pass ``lgas`` (from ``allowed_lgas_of(user)``) to additionally restrict an
    LGA-scoped account to its own LGA(s); ``lga_col`` names the LGA column on the
    primary table (``lga`` for mda_households, ``lga_name`` for analytics/geo).
    """
    col = f"{alias}.project_id" if alias else "project_id"
    filters.insert(0, f"{col} = :pid")
    if date_filter:
        rcv = f"{alias}.{date_col}" if alias else date_col
        # COALESCE makes this a no-op when campaign_start_date is NULL,
        # so existing projects without the override behave exactly as before.
        filters.insert(1, (
            f"(({rcv} AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date "
            f">= COALESCE("
            f"(SELECT campaign_start_date FROM geo_projects WHERE id = :pid),"
            f" '1900-01-01'::date))"
        ))
    params["pid"] = pid
    lcol = f"{alias}.{lga_col}" if alias else lga_col
    return "WHERE " + " AND ".join(filters) + _lga_and(lgas, lcol, params)


async def deny_lga_scoped(user: "Optional[User]" = Depends(get_current_user_optional)) -> None:
    """Dependency guard for cross-LGA / cross-round endpoints that can't be
    meaningfully narrowed to a single LGA (round-over-round comparisons, etc.).
    An LGA-scoped account is refused (403); admins, state-only users and the
    public pass through unchanged."""
    if allowed_lgas_of(user) is not None:
        raise HTTPException(status_code=403, detail="This view isn't available for LGA-scoped accounts.")


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/mda/upload
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_mda(
    file: UploadFile = File(...),
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _super: User = Depends(require_superadmin),
):
    """
    Replace MDA data for the active project with the uploaded Excel workbook.
    Other projects (other rounds / other states) are untouched.
    Parses Forms sheet (households) and Repeat-group_indv sheet (individuals).
    Returns QC flag summary counts.
    """
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".csv")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .csv")

    raw_bytes = await file.read()

    # ── Read input ──────────────────────────────────────────────────────────
    # CSV path treats the whole file as the Forms sheet (households only — no
    # individual repeat-group). For full ingest with individuals, the user
    # should upload an .xlsx with Forms + Repeat-group_indv sheets.
    wb = None  # only set for xlsx path; used later for individuals parse
    if fname.endswith(".csv"):
        import csv as _csv
        try:
            text_blob = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text_blob = raw_bytes.decode("latin-1")
        reader = _csv.reader(io.StringIO(text_blob))
        rows_forms = [tuple(r) for r in reader]
    else:
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(raw_bytes), read_only=True, data_only=True
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Cannot open workbook: {exc}")

        sheet_names = wb.sheetnames
        if "Forms" not in sheet_names:
            raise HTTPException(status_code=400, detail="Workbook must contain a 'Forms' sheet")

        # ── Parse Forms (households) ─────────────────────────────────────────────
        ws_forms = wb["Forms"]
        rows_forms = list(ws_forms.iter_rows(values_only=True))

    households: List[Dict[str, Any]] = []
    seen_formids: Dict[str, int] = {}   # formid → first occurrence index
    # Row 0 is the header — skip it
    for row_idx, row in enumerate(rows_forms[1:], start=1):
        def cell(idx):
            try:
                return row[idx]
            except IndexError:
                return None

        formid = _str(cell(1))
        if not formid:
            continue  # skip blank rows

        teamcode       = _str(cell(2))
        data_type      = _str(cell(3))
        trt_day        = _str(cell(4))
        data_entry_persons = _str(cell(5))
        phone_number_data  = _str(cell(6))
        admin3_code    = _str(cell(8))
        admin5_code    = _str(cell(10))
        consent_trt    = _str(cell(11))
        reasons_for_refusal        = _str(cell(12))
        others_reasons_for_refusal = _str(cell(13))
        gps_raw        = _str(cell(14))
        date_trt_raw   = _parse_date(cell(15))
        hh_num         = _str(cell(16))
        lga            = _str(cell(22))
        hh_seq         = _str(cell(27))
        serial_number_hh_id = _str(cell(28))
        number_of_treated   = _int_safe(cell(30))
        housemarking_code   = _str(cell(31))
        completed_time = _parse_dt(cell(32))
        started_time   = _parse_dt(cell(33))
        username       = _str(cell(34))
        received_on    = _parse_dt(cell(35))
        check_treatment_date_raw = _parse_date(cell(21))
        hq_user_val = _str(cell(37))

        # Normalize RA name
        name_norm = data_entry_persons.lower().strip() if data_entry_persons else None
        ra_key = f"{name_norm}|{phone_number_data}" if name_norm and phone_number_data else None

        # GPS
        lat, lon, _alt, accuracy = _parse_gps(gps_raw)

        # Durations
        form_duration_min = None
        if started_time and completed_time:
            delta = (completed_time - started_time).total_seconds()
            form_duration_min = round(delta / 60.0, 2)

        sync_lag_hours = None
        if completed_time and received_on:
            delta = (received_on - completed_time).total_seconds()
            sync_lag_hours = round(delta / 3600.0, 2)

        # QC flags
        flag_gps_poor_accuracy = bool(accuracy is not None and accuracy > 20)
        flag_gps_zero = bool(
            lat is not None and lon is not None and lat == 0.0 and lon == 0.0
        )

        # After-hours: convert UTC to UTC+1 before checking
        flag_after_hours = False
        if started_time:
            local_hour = (started_time + timedelta(hours=1)).hour
            flag_after_hours = local_hour < 6 or local_hour >= 19

        # Fast-form: <3 min AND a real visit was actually attempted. Refusals
        # (consent_trt='0') and N/A entries (consent_trt='2') finish quickly
        # for legitimate reasons and are excluded (operator agreement
        # 2026-07-05). The CommCare sync path also excludes empty households
        # via consent_survey.num_reside='0' — this Excel workbook format
        # doesn't carry num_reside as a mapped cell, so it applies only the
        # consent_trt guard here.
        _real_visit_attempted = consent_trt not in ("0", "2")
        flag_fast_form = bool(
            form_duration_min is not None and form_duration_min < 3
            and _real_visit_attempted
        )
        flag_slow_form = bool(form_duration_min is not None and form_duration_min > 60)
        flag_sync_lag  = bool(sync_lag_hours is not None and sync_lag_hours > 48)
        flag_refusal   = consent_trt == "0"

        # Duplicate detection (set to True for subsequent occurrences)
        flag_duplicate = formid in seen_formids
        if not flag_duplicate:
            seen_formids[formid] = row_idx

        # Geometry WKT
        geom_wkt = None
        if lat is not None and lon is not None and not flag_gps_zero:
            geom_wkt = f"SRID=4326;POINT({lon} {lat})"

        households.append({
            "formid": formid,
            "username": username,
            "teamcode": teamcode,
            "data_type": data_type,
            "data_entry_persons": data_entry_persons,
            "data_entry_persons_norm": name_norm,
            "phone_number_data": phone_number_data,
            "ra_key": ra_key,
            "lga": lga,
            "admin3_code": admin3_code,
            "admin5_code": admin5_code,
            "trt_day": trt_day,
            "date_trt": date_trt_raw,
            "consent_trt": consent_trt,
            "reasons_for_refusal": reasons_for_refusal,
            "others_reasons_for_refusal": others_reasons_for_refusal,
            "hh_num": hh_num,
            "hh_seq": hh_seq,
            "serial_number_hh_id": serial_number_hh_id,
            "number_of_treated": number_of_treated,
            "housemarking_code": housemarking_code,
            "gps_raw": gps_raw,
            "latitude": lat,
            "longitude": lon,
            "gps_accuracy": accuracy,
            "geom_wkt": geom_wkt,
            "started_time": started_time.isoformat() if started_time else None,
            "completed_time": completed_time.isoformat() if completed_time else None,
            "received_on": received_on.isoformat() if received_on else None,
            "form_duration_min": form_duration_min,
            "sync_lag_hours": sync_lag_hours,
            "flag_duplicate": flag_duplicate,
            "flag_duplicate_gps": False,          # set by SQL after insert
            "flag_gps_outside_lga": False,        # set by SQL after insert
            "flag_gps_outside_ward": False,       # set by SQL after insert
            "flag_gps_outside_state": False,      # set by SQL after insert
            "flag_gps_poor_accuracy": flag_gps_poor_accuracy,
            "flag_gps_zero": flag_gps_zero,
            "flag_after_hours": flag_after_hours,
            "flag_fast_form": flag_fast_form,
            "flag_slow_form": flag_slow_form,
            "flag_sync_lag": flag_sync_lag,
            "flag_refusal": flag_refusal,
            "check_treatment_date": check_treatment_date_raw,
            "hq_user": hq_user_val,
        })

    # ── Parse Repeat-group_indv (individuals) ────────────────────────────────
    # CSV uploads can't carry the individuals repeat-group — they're a separate
    # sheet in the canonical SARMAAN workbook. CSV path skips this section so
    # households still upsert; individuals stay at whatever's currently in the DB.
    individuals: List[Dict[str, Any]] = []
    valid_hh_formids = {h["formid"] for h in households if not h["flag_duplicate"]}

    if wb is not None and "Repeat- group_indv" in wb.sheetnames:
        ws_indv = wb["Repeat- group_indv"]
        rows_indv = list(ws_indv.iter_rows(values_only=True))
        for row in rows_indv[1:]:
            def icell(idx):
                try:
                    return row[idx]
                except IndexError:
                    return None

            hh_formid   = _str(icell(17))
            if not hh_formid:
                continue

            mother_name       = _str(icell(3))
            child_name        = _str(icell(4))
            dob               = _parse_date(icell(5))
            dob_checknote     = _str(icell(6))
            sex               = _str(icell(7))
            height_cm         = _str(icell(8))
            treatment_status  = _str(icell(9))
            not_treated       = _str(icell(10))
            vomit_spill_azt   = _str(icell(11))
            child_id_r2       = _str(icell(12))
            respondent_hh_id  = _str(icell(16))
            age_in_months     = _int_safe(icell(18))
            individual_id     = _str(icell(20))

            flag_orphan = hh_formid not in valid_hh_formids

            individuals.append({
                "hh_formid": hh_formid,
                "mother_name": mother_name,
                "child_name": child_name,
                "dob": dob,
                "dob_checknote": dob_checknote,
                "sex": sex,
                "height_cm": height_cm,
                "age_in_months": age_in_months,
                "treatment_status": treatment_status,
                "not_treated": not_treated,
                "vomit_spill_azt": vomit_spill_azt,
                "child_id_r2": child_id_r2,
                "respondent_hh_id": respondent_hh_id,
                "individual_id": individual_id,
                "flag_orphan": flag_orphan,
            })

    if wb is not None:
        wb.close()

    # ── Bulk insert via psycopg2 ─────────────────────────────────────────────
    conn = _get_sync_conn()
    try:
        cur = conn.cursor()

        # Resolve the boundary-owning project for the same state (lowest-id project
        # in the same state). For Sokoto that's R4 (id=1) — boundaries are state-level.
        cur.execute("""
            SELECT MIN(p2.id) FROM geo_projects p1
            JOIN geo_projects p2 ON p2.state_name = p1.state_name
            WHERE p1.id = %s
        """, (pid,))
        boundary_pid = (cur.fetchone() or [pid])[0] or pid

        # Delete existing data — scoped to the active project only
        cur.execute("DELETE FROM mda_individuals WHERE project_id = %s", (pid,))
        cur.execute("DELETE FROM mda_households  WHERE project_id = %s", (pid,))

        # Insert households
        if households:
            hh_cols = [
                "project_id",
                "formid", "username", "teamcode", "data_type",
                "data_entry_persons", "data_entry_persons_norm",
                "phone_number_data", "ra_key", "lga",
                "admin3_code", "admin5_code", "trt_day", "date_trt",
                "consent_trt", "reasons_for_refusal", "others_reasons_for_refusal",
                "hh_num", "hh_seq", "serial_number_hh_id", "number_of_treated",
                "housemarking_code", "gps_raw", "latitude", "longitude",
                "gps_accuracy", "geom",
                "started_time", "completed_time", "received_on",
                "form_duration_min", "sync_lag_hours",
                "flag_duplicate", "flag_duplicate_gps",
                "flag_gps_outside_lga", "flag_gps_outside_ward", "flag_gps_outside_state",
                "flag_gps_poor_accuracy", "flag_gps_zero", "flag_after_hours",
                "flag_fast_form", "flag_slow_form", "flag_sync_lag", "flag_refusal",
                "check_treatment_date", "hq_user",
                "uploaded_at",
            ]

            hh_values = []
            now_str = datetime.utcnow().isoformat()
            for h in households:
                geom_val = h["geom_wkt"]  # WKT string or None
                hh_values.append((
                    pid,
                    h["formid"], h["username"], h["teamcode"], h["data_type"],
                    h["data_entry_persons"], h["data_entry_persons_norm"],
                    h["phone_number_data"], h["ra_key"], h["lga"],
                    h["admin3_code"], h["admin5_code"], h["trt_day"], h["date_trt"],
                    h["consent_trt"], h["reasons_for_refusal"], h["others_reasons_for_refusal"],
                    h["hh_num"], h["hh_seq"], h["serial_number_hh_id"], h["number_of_treated"],
                    h["housemarking_code"], h["gps_raw"], h["latitude"], h["longitude"],
                    h["gps_accuracy"], geom_val,
                    h["started_time"], h["completed_time"], h["received_on"],
                    h["form_duration_min"], h["sync_lag_hours"],
                    # flags — must match hh_cols order exactly
                    h["flag_duplicate"], h["flag_duplicate_gps"],
                    h["flag_gps_outside_lga"], h["flag_gps_outside_ward"], h["flag_gps_outside_state"],
                    h["flag_gps_poor_accuracy"], h["flag_gps_zero"], h["flag_after_hours"],
                    h["flag_fast_form"], h["flag_slow_form"], h["flag_sync_lag"], h["flag_refusal"],
                    h["check_treatment_date"], h["hq_user"],
                    now_str,
                ))

            col_str = ", ".join(hh_cols)
            # Build placeholders; geom column uses ST_GeomFromEWKT
            placeholders = []
            for col in hh_cols:
                if col == "geom":
                    placeholders.append("ST_GeomFromEWKT(%s)")
                else:
                    placeholders.append("%s")
            ph_str = ", ".join(placeholders)

            psycopg2.extras.execute_values(
                cur,
                f"INSERT INTO mda_households ({col_str}) VALUES %s",
                hh_values,
                template=f"({ph_str})",
                page_size=500,
            )

        # Insert individuals
        if individuals:
            indv_cols = [
                "project_id",
                "hh_formid", "mother_name", "child_name", "dob", "dob_checknote",
                "sex", "height_cm", "age_in_months", "treatment_status",
                "not_treated", "vomit_spill_azt", "child_id_r2",
                "respondent_hh_id", "individual_id", "flag_orphan", "uploaded_at",
            ]
            now_str = datetime.utcnow().isoformat()
            indv_values = [
                (
                    pid,
                    i["hh_formid"], i["mother_name"], i["child_name"],
                    i["dob"], i["dob_checknote"], i["sex"], i["height_cm"],
                    i["age_in_months"], i["treatment_status"], i["not_treated"],
                    i["vomit_spill_azt"], i["child_id_r2"], i["respondent_hh_id"],
                    i["individual_id"], i["flag_orphan"], now_str,
                )
                for i in individuals
            ]
            col_str = ", ".join(indv_cols)
            psycopg2.extras.execute_values(
                cur,
                f"INSERT INTO mda_individuals ({col_str}) VALUES %s",
                indv_values,
                page_size=1000,
            )

        # GPS outside stated LGA polygon (scoped to this project's households +
        # the state's canonical boundary project)
        cur.execute("""
            UPDATE mda_households h
            SET flag_gps_outside_lga = TRUE
            WHERE h.project_id = %s
              AND h.geom IS NOT NULL AND h.flag_gps_zero = FALSE
              AND NOT EXISTS (
                  SELECT 1 FROM lgas l
                  WHERE l.project_id = %s
                    AND UPPER(TRIM(l.lga_name)) = UPPER(TRIM(h.lga))
                    AND ST_Within(h.geom, l.geom)
              )
        """, (pid, boundary_pid))

        # GPS outside any state LGA polygon
        cur.execute("""
            UPDATE mda_households h
            SET flag_gps_outside_state = TRUE
            WHERE h.project_id = %s
              AND h.geom IS NOT NULL AND h.flag_gps_zero = FALSE
              AND NOT EXISTS (
                  SELECT 1 FROM lgas l
                  WHERE l.project_id = %s
                    AND ST_Within(h.geom, l.geom)
              )
        """, (pid, boundary_pid))

        # GPS outside any ward polygon
        cur.execute("""
            UPDATE mda_households h
            SET flag_gps_outside_ward = TRUE
            WHERE h.project_id = %s
              AND h.geom IS NOT NULL AND h.flag_gps_zero = FALSE
              AND NOT EXISTS (
                  SELECT 1 FROM wards w
                  WHERE w.project_id = %s
                    AND ST_Within(h.geom, w.geom)
              )
        """, (pid, boundary_pid))

        # Duplicate GPS coordinates within this project
        cur.execute("""
            UPDATE mda_households h
            SET flag_duplicate_gps = TRUE
            WHERE h.project_id = %s
              AND h.latitude IS NOT NULL AND h.longitude IS NOT NULL
              AND h.flag_gps_zero = FALSE
              AND EXISTS (
                  SELECT 1 FROM mda_households h2
                  WHERE h2.project_id = %s
                    AND h2.id != h.id
                    AND h2.latitude = h.latitude
                    AND h2.longitude = h.longitude
              )
        """, (pid, pid))

        # Spatially join each household GPS point to the ward boundary → populate ward_name
        cur.execute("""
            UPDATE mda_households h
            SET ward_name = w.ward_name
            FROM wards w
            WHERE h.project_id = %s
              AND w.project_id = %s
              AND ST_Within(h.geom, w.geom)
              AND h.geom IS NOT NULL
              AND h.flag_gps_zero = FALSE
        """, (pid, boundary_pid))

        # Normalise lga names to match shapefile Title Case (e.g. TAMBUWAL → Tambuwal)
        cur.execute("""
            UPDATE mda_households h
            SET lga = w.lga_name
            FROM (SELECT DISTINCT lga_name FROM wards WHERE project_id = %s) w
            WHERE h.project_id = %s
              AND UPPER(TRIM(h.lga)) = UPPER(TRIM(w.lga_name))
              AND h.lga <> w.lga_name
        """, (boundary_pid, pid))

        conn.commit()
    except Exception as exc:
        conn.rollback()
        logger.exception("MDA bulk insert failed")
        raise HTTPException(status_code=500, detail=f"Database error: {exc}")
    finally:
        conn.close()

    # ── Compute QC flag counts to return ────────────────────────────────────
    qc_flags = {
        "duplicates":        sum(1 for h in households if h["flag_duplicate"]),
        "gps_poor_accuracy": sum(1 for h in households if h["flag_gps_poor_accuracy"]),
        "gps_zero":          sum(1 for h in households if h["flag_gps_zero"]),
        "after_hours":       sum(1 for h in households if h["flag_after_hours"]),
        "fast_forms":        sum(1 for h in households if h["flag_fast_form"]),
        "slow_forms":        sum(1 for h in households if h["flag_slow_form"]),
        "sync_lag":          sum(1 for h in households if h["flag_sync_lag"]),
        "refusals":          sum(1 for h in households if h["flag_refusal"]),
        "orphan_individuals":sum(1 for i in individuals if i["flag_orphan"]),
    }

    return {
        "households": len(households),
        "individuals": len(individuals),
        "qc_flags": qc_flags,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/landing-stats  — public, drives the login page hero
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/landing-stats")
async def landing_stats(db: AsyncSession = Depends(get_db)):
    """Lightweight public snapshot for the login page.

    No auth required. Returns the active round's high-level KPIs so the
    landing-page tiles can show live progress instead of static placeholders.
    """
    # Pick the currently-active project; fall back to the most recent one.
    proj = await db.execute(text("""
        SELECT id, state_name, round_number, campaign_start_date, campaign_end_date,
               COALESCE(campaign_paused, FALSE) AS campaign_paused
        FROM geo_projects
        ORDER BY is_active DESC, round_number DESC NULLS LAST, id DESC
        LIMIT 1
    """))
    p = proj.fetchone()
    if not p:
        return {
            "active_round_label": "—",
            "is_live": False,
            "total_treated": 0,
            "baseline_total": 0,
            "coverage_pct": 0,
            "lgas_in_scope": 0,
            "wards_in_scope": 0,
            "days_active": 0,
            "planned_duration_days": None,
        }

    pid = p[0]
    label = (
        f"{p[1]} Round {p[2]}" if p[1] and p[2] is not None else f"Project #{pid}"
    )
    # "Campaign Days" = the official planned duration from the configured
    # campaign window (end − start + 1), NOT the count of distinct submission
    # days, which is inflated by pre-campaign test forms and a late/out-of-window
    # trickle. Falls back to None when no window is set (the UI then shows
    # days_active instead).
    start_d, end_d = p[3], p[4]
    planned_days = (end_d - start_d).days + 1 if (start_d and end_d) else None
    # "live" means the campaign is actually running - started and not yet ended.
    # A round that is only the default/shown one but hasn't started, or has
    # ended, is NOT live.
    _today = datetime.utcnow().date()
    _paused = bool(p[5])
    is_live = bool(start_d and start_d <= _today and (end_d is None or end_d >= _today) and not _paused)

    # Boundary tables are typically loaded under a single project per state
    # (R4 holds Sokoto's polygons; R5 reuses them) — count distinct codes
    # across every project that shares this round's state so the tiles still
    # reflect the real scope even when the active round didn't carry its own
    # boundary import.
    state_name = p[1]
    res = await db.execute(text("""
        SELECT
          (SELECT COALESCE(SUM(number_of_treated), 0) FROM mda_households WHERE project_id = :pid) AS treated,
          (SELECT COALESCE(SUM(total_treated),   0) FROM mda_baseline    WHERE project_id = :pid) AS baseline,
          (SELECT COUNT(DISTINCT (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)
                   FROM mda_households WHERE project_id = :pid)                                   AS days_active,
          (SELECT COUNT(DISTINCT lgacode) FROM lgas
            WHERE project_id IN (SELECT id FROM geo_projects WHERE state_name = :state OR :state IS NULL))  AS lgas_in_scope,
          (SELECT COUNT(DISTINCT wardcode) FROM wards
            WHERE project_id IN (SELECT id FROM geo_projects WHERE state_name = :state OR :state IS NULL)) AS wards_in_scope
    """), {"pid": pid, "state": state_name})
    row = res.fetchone()
    treated  = int(row.treated or 0)
    baseline = int(row.baseline or 0)
    return {
        "active_round_label": label,
        "is_live":         is_live,
        "total_treated":   treated,
        "baseline_total":  baseline,
        "coverage_pct":    round(100.0 * treated / baseline, 1) if baseline > 0 else 0,
        "lgas_in_scope":   int(row.lgas_in_scope or 0),
        "wards_in_scope":  int(row.wards_in_scope or 0),
        "days_active":     int(row.days_active or 0),
        "planned_duration_days": planned_days,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/rounds/summary — per-round comparison rows
# Used by the Campaign Trends "Round-over-Round Comparison" table. Returns one
# row per project (R4, R5, …) with period, targeted, administered, coverage,
# refusals, and QC-flag counts. The frontend computes Δ vs prev round.
# Public (no auth) so it can sit alongside the other Overview/Trends widgets.
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/rounds/summary")
async def rounds_summary(
    project_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Per-round summary. When ``project_id`` is supplied, the query scopes
    to that project's state — so switching to "Kano Round 1" only returns
    Kano rounds, not Sokoto. With no ``project_id`` it returns every round
    of every state (the original cross-state view). An LGA-scoped account sees
    each round's totals for its own LGA only."""
    # Every per-project subquery filters received_on >= the project's official
    # campaign_start_date (when set) so pre-campaign test submissions are
    # excluded from period dates, totals, refusal counts and QC totals.
    # COALESCE keeps it a no-op for projects without a configured start.
    params: dict = {}
    lgas = allowed_lgas_of(_u)
    lga_h = _lga_and(lgas, "h.lga", params)   # baked into every household subquery
    lga_b = _lga_and(lgas, "b.lga", params)   # for the baseline subquery
    DATE_CLAUSE = f"""
        AND (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
            >= COALESCE(p.campaign_start_date, '1900-01-01'::date){lga_h}
    """
    state_filter_sql = ""
    if project_id is not None:
        state_filter_sql = """
        WHERE p.state_name = (SELECT state_name FROM geo_projects WHERE id = :pid)
        """
        params["pid"] = project_id
    res = await db.execute(text(f"""
        WITH per_project AS (
            SELECT
              p.id, p.state_name, p.round_number, p.is_active,
              (SELECT COALESCE(SUM(total_treated), 0)
                 FROM mda_baseline b WHERE b.project_id = p.id{lga_b}) AS targeted,
              (SELECT COALESCE(SUM(number_of_treated), 0)
                 FROM mda_households h WHERE h.project_id = p.id {DATE_CLAUSE}) AS administered,
              (SELECT MIN((received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)
                 FROM mda_households h WHERE h.project_id = p.id {DATE_CLAUSE}) AS period_start,
              (SELECT MAX((received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)
                 FROM mda_households h WHERE h.project_id = p.id {DATE_CLAUSE}) AS period_end,
              (SELECT COUNT(*) FROM mda_households h
                 WHERE h.project_id = p.id AND h.flag_refusal = TRUE {DATE_CLAUSE}) AS refusals,
              (SELECT COUNT(*) FROM mda_households h WHERE h.project_id = p.id {DATE_CLAUSE}) AS total_forms,
              (SELECT COALESCE(SUM(
                  (CASE WHEN h.flag_gps_outside_lga    THEN 1 ELSE 0 END)
                + (CASE WHEN h.flag_duplicate_gps      THEN 1 ELSE 0 END)
                + (CASE WHEN h.flag_gps_poor_accuracy THEN 1 ELSE 0 END)
                + (CASE WHEN h.flag_after_hours        THEN 1 ELSE 0 END)
                + (CASE WHEN h.flag_fast_form          THEN 1 ELSE 0 END)
                + (CASE WHEN h.flag_slow_form          THEN 1 ELSE 0 END)
              ),0) FROM mda_households h WHERE h.project_id = p.id {DATE_CLAUSE}) AS qc_flags
            FROM geo_projects p
            {state_filter_sql}
        )
        SELECT id, state_name, round_number, is_active,
               targeted, administered, period_start, period_end,
               refusals, total_forms, qc_flags
        FROM per_project
        ORDER BY round_number NULLS LAST, id
    """), params)
    out = []
    for r in res.fetchall():
        targeted = int(r.targeted or 0)
        administered = int(r.administered or 0)
        total_forms = int(r.total_forms or 0)
        refusals = int(r.refusals or 0)
        out.append({
            "project_id":    r.id,
            "round_number":  r.round_number,
            "round_label":   (f"Round {r.round_number}" if r.round_number is not None else f"Project #{r.id}"),
            "state_name":    r.state_name,
            "is_active":     bool(r.is_active),
            "period_start":  r.period_start.isoformat() if r.period_start else None,
            "period_end":    r.period_end.isoformat()   if r.period_end   else None,
            "targeted":      targeted,
            "administered":  administered,
            "coverage_pct":  round(100.0 * administered / targeted, 1) if targeted > 0 else 0.0,
            "refusals":      refusals,
            "refusal_pct":   round(100.0 * refusals / total_forms, 2) if total_forms > 0 else 0.0,
            "total_forms":   total_forms,
            "qc_flags":      int(r.qc_flags or 0),
        })
    return out


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/trends/daily-by-round — daily series per round
# Returns one row per (project, day_index). day_index = 0 on each round's start
# date, so R4 day-1 lines up with R5 day-1 on the chart even when the calendar
# dates differ. Used by the Campaign Trends "Daily Form Submissions" and
# "Daily Cumulative Administration" cards so R4 and R5 plot as parallel lines.
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/trends/daily-by-round")
async def trends_daily_by_round(
    project_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Daily series per round. When ``project_id`` is supplied, the response
    is scoped to that project's state — so switching to Kano returns only
    Kano rounds, not Sokoto. An LGA-scoped account sees only its own LGA's
    daily series across rounds."""
    # Daily series is also bounded by the project's campaign_start_date so the
    # Day-1 alignment used by the Campaign Trends charts matches the official
    # campaign start (e.g. R5 begins Day-1 on 19 May, not the 18 May test).
    state_filter_sql = ""
    params: dict = {}
    if project_id is not None:
        state_filter_sql = "AND p.state_name = (SELECT state_name FROM geo_projects WHERE id = :pid)"
        params["pid"] = project_id
    lga_h = _lga_and(allowed_lgas_of(_u), "h.lga", params)
    res = await db.execute(text(f"""
        WITH per_day AS (
            SELECT h.project_id,
                   (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date AS day,
                   COUNT(*) AS forms,
                   COALESCE(SUM(h.number_of_treated), 0) AS treated
            FROM mda_households h
            JOIN geo_projects p ON p.id = h.project_id
            WHERE h.received_on IS NOT NULL
              AND (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
                  >= COALESCE(p.campaign_start_date, '1900-01-01'::date){lga_h}
              {state_filter_sql}
            GROUP BY h.project_id, (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
        ),
        per_project AS (
            SELECT project_id, MIN(day) AS start_day
            FROM per_day GROUP BY project_id
        )
        SELECT p.id AS project_id, p.round_number, p.state_name,
               d.day,
               (d.day - pp.start_day) AS day_index,
               d.forms, d.treated
        FROM geo_projects p
        JOIN per_day d ON d.project_id = p.id
        JOIN per_project pp ON pp.project_id = p.id
        ORDER BY p.round_number NULLS LAST, p.id, d.day
    """), params)
    by_round: dict = {}
    for r in res.fetchall():
        key = r.round_number if r.round_number is not None else f"P{r.project_id}"
        bucket = by_round.setdefault(key, {
            "project_id": r.project_id,
            "round_number": r.round_number,
            "round_label": f"Round {r.round_number}" if r.round_number is not None else f"Project #{r.project_id}",
            "state_name": r.state_name,
            "days": [],
        })
        bucket["days"].append({
            "day":       r.day.isoformat() if r.day else None,
            "day_index": int(r.day_index or 0),
            "forms":     int(r.forms or 0),
            "treated":   int(r.treated or 0),
        })
    return list(by_round.values())


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/rounds/lga-compare — per-LGA R4 vs R5 (and any other rounds)
# Returns one row per LGA with each round's forms, treated, and coverage_pct
# alongside the delta from the previous round so the Campaign Trends page can
# highlight LGAs that improved or regressed between rounds.
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/rounds/lga-compare")
async def rounds_lga_compare(
    project_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Per-LGA cross-round comparison. When ``project_id`` is supplied, the
    query is scoped to that project's state — so switching to Kano returns
    only Kano LGAs, not Sokoto. An LGA-scoped account sees only its own LGA's
    row(s)."""
    params: dict = {}
    lgas = allowed_lgas_of(_u)
    lga_b = _lga_and(lgas, "b.lga", params)
    lga_h = _lga_and(lgas, "h.lga", params)
    # per_round always carries a WHERE so the state + LGA clauses append as AND.
    state_filter_pr = "WHERE TRUE"
    state_filter_tr = ""  # in treated CTE   (p is geo_projects, joined)
    if project_id is not None:
        state_filter_pr += " AND p.state_name = (SELECT state_name FROM geo_projects WHERE id = :pid)"
        state_filter_tr = "AND p.state_name = (SELECT state_name FROM geo_projects WHERE id = :pid)"
        params["pid"] = project_id
    state_filter_pr += lga_b
    res = await db.execute(text(f"""
        WITH per_round AS (
            SELECT p.id AS project_id,
                   p.round_number,
                   INITCAP(TRIM(b.lga)) AS lga,
                   COALESCE(SUM(b.total_treated), 0) AS baseline_total
            FROM geo_projects p
            JOIN mda_baseline b ON b.project_id = p.id
            {state_filter_pr}
            GROUP BY p.id, p.round_number, INITCAP(TRIM(b.lga))
        ),
        treated AS (
            SELECT h.project_id, INITCAP(TRIM(h.lga)) AS lga,
                   COUNT(*) AS forms,
                   COALESCE(SUM(h.number_of_treated), 0) AS treated
            FROM mda_households h
            JOIN geo_projects p ON p.id = h.project_id
            WHERE h.lga IS NOT NULL
              AND (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
                  >= COALESCE(p.campaign_start_date, '1900-01-01'::date){lga_h}
              {state_filter_tr}
            GROUP BY h.project_id, INITCAP(TRIM(h.lga))
        )
        SELECT pr.project_id, pr.round_number, pr.lga, pr.baseline_total,
               COALESCE(t.forms, 0)   AS forms,
               COALESCE(t.treated, 0) AS treated
        FROM per_round pr
        LEFT JOIN treated t ON t.project_id = pr.project_id AND t.lga = pr.lga
        ORDER BY pr.lga, pr.round_number NULLS LAST
    """), params)
    rounds_ordered: list = []
    by_lga: dict = {}
    seen_rounds = []
    for r in res.fetchall():
        rn = r.round_number
        if rn not in seen_rounds:
            seen_rounds.append(rn)
        row = by_lga.setdefault(r.lga, {"lga": r.lga, "rounds": {}})
        bl = int(r.baseline_total or 0)
        tr = int(r.treated or 0)
        row["rounds"][rn] = {
            "round_number": rn,
            "baseline":     bl,
            "forms":        int(r.forms or 0),
            "treated":      tr,
            "coverage_pct": round(100.0 * tr / bl, 1) if bl > 0 else 0.0,
        }
    rounds_meta = [{"round_number": rn} for rn in sorted([x for x in seen_rounds if x is not None])]
    # The round the caller is actually looking at, so the client shows THAT
    # round's per-LGA progress instead of a hardcoded one. When scoped to a
    # project, it's that project's round; otherwise the latest round present.
    active_round = None
    if project_id is not None:
        active_round = (await db.execute(
            text("SELECT round_number FROM geo_projects WHERE id = :pid"),
            {"pid": project_id},
        )).scalar()
    if active_round is None and rounds_meta:
        active_round = rounds_meta[-1]["round_number"]
    return {"rounds": rounds_meta, "active_round": active_round, "lgas": list(by_lga.values())}


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/system/counts — admin diagnostics for the System Status page
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/system/counts")
async def system_counts(
    db: AsyncSession = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Single round-trip for the System Status tiles.

    Household / individual / baseline counts are scoped to the active round
    (the same scope the dashboard renders). Boundary counts are state-wide
    since polygons are loaded once per state and reused across rounds.
    """
    proj = await db.execute(text("""
        SELECT id, state_name FROM geo_projects
        ORDER BY is_active DESC, round_number DESC NULLS LAST, id DESC
        LIMIT 1
    """))
    p = proj.fetchone()
    pid = p[0] if p else None
    state = p[1] if p else None

    res = await db.execute(text("""
        SELECT
          (SELECT COUNT(*) FROM mda_households   WHERE project_id = :pid)                       AS households,
          (SELECT COUNT(*) FROM mda_individuals  WHERE project_id = :pid)                       AS individuals,
          (SELECT COUNT(*) FROM mda_baseline     WHERE project_id = :pid)                       AS baseline_records,
          (SELECT COUNT(DISTINCT lgacode)  FROM lgas
            WHERE project_id IN (SELECT id FROM geo_projects WHERE state_name = :state OR :state IS NULL))   AS lgas,
          (SELECT COUNT(DISTINCT wardcode) FROM wards
            WHERE project_id IN (SELECT id FROM geo_projects WHERE state_name = :state OR :state IS NULL))   AS wards,
          (SELECT COUNT(*)                 FROM settlements
            WHERE project_id IN (SELECT id FROM geo_projects WHERE state_name = :state OR :state IS NULL))   AS settlements
    """), {"pid": pid, "state": state})
    row = res.fetchone()
    return {
        "active_project_id": pid,
        "households":        int(row.households or 0),
        "individuals":       int(row.individuals or 0),
        "baseline_records":  int(row.baseline_records or 0),
        "lgas":              int(row.lgas or 0),
        "wards":             int(row.wards or 0),
        "settlements":       int(row.settlements or 0),
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/qc/summary
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/qc/summary")
async def qc_summary(
    lga:       Optional[str] = None,
    ward:      Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    params: dict = {}
    filters: list = []
    if lga:
        filters.append("lga = :lga")
        params["lga"] = lga
    if ward:
        filters.append("REPLACE(LOWER(TRIM(ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')")
        params["ward"] = ward
    # Date filter / "days active" both key off received_on — the timestamp at
    # which CommCare HQ accepted the form. Field workers can backfill date_trt
    # / check_treatment_date inconsistently, so received_on is the only column
    # that reflects what the platform actually has data for.
    if date_from: filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date >= :date_from"); params["date_from"] = _as_date(date_from)
    if date_to:   filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date <= :date_to");   params["date_to"]   = _as_date(date_to)
    where = _scoped_where(pid, filters, params, lgas=allowed_lgas_of(_u))
    result = await db.execute(text(f"""
        SELECT
          COUNT(*) AS total_forms,
          SUM(CASE WHEN flag_duplicate THEN 1 ELSE 0 END) AS duplicates,
          SUM(CASE WHEN flag_duplicate_gps THEN 1 ELSE 0 END) AS duplicate_gps,
          SUM(CASE WHEN flag_gps_outside_lga THEN 1 ELSE 0 END) AS gps_outside_lga,
          SUM(CASE WHEN flag_gps_outside_ward THEN 1 ELSE 0 END) AS gps_outside_ward,
          SUM(CASE WHEN flag_gps_outside_state THEN 1 ELSE 0 END) AS gps_outside_state,
          SUM(CASE WHEN flag_gps_poor_accuracy THEN 1 ELSE 0 END) AS gps_poor_accuracy,
          SUM(CASE WHEN flag_gps_zero THEN 1 ELSE 0 END) AS gps_zero,
          SUM(CASE WHEN flag_after_hours THEN 1 ELSE 0 END) AS after_hours,
          SUM(CASE WHEN flag_fast_form THEN 1 ELSE 0 END) AS fast_forms,
          SUM(CASE WHEN flag_slow_form THEN 1 ELSE 0 END) AS slow_forms,
          SUM(CASE WHEN flag_sync_lag THEN 1 ELSE 0 END) AS sync_lag,
          SUM(CASE WHEN flag_refusal THEN 1 ELSE 0 END) AS refusals,
          COUNT(*) FILTER (WHERE flag_refusal) * 100.0 / NULLIF(COUNT(*), 0) AS refusal_pct,
          -- forms_with_error: at least one *real* error flag (excluding
          -- refusal which is a campaign outcome, not a data-quality issue).
          SUM(CASE WHEN (
              flag_fast_form OR flag_slow_form OR flag_after_hours
              OR flag_gps_outside_lga OR flag_gps_poor_accuracy
              OR flag_duplicate_gps OR flag_duplicate OR flag_gps_zero
          ) THEN 1 ELSE 0 END) AS forms_with_error,
          COUNT(DISTINCT (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date) AS days_active,
          COUNT(DISTINCT lga) AS lgas_covered,
          COUNT(DISTINCT ra_key) AS ra_count,
          (SELECT COUNT(*) FROM mda_individuals WHERE project_id = :pid) AS total_individuals,
          (SELECT MIN(uploaded_at) FROM mda_households WHERE project_id = :pid) AS data_as_of,
          -- Form-duration stats. completed_time and started_time are stored
          -- timezone-aware (UTC); duration = completed - started in minutes,
          -- already precomputed into form_duration_min at ingest. Filter
          -- non-null + clip the long tail (>= 6 hours = device left open) so
          -- the mean is dominated by real visits, not abandoned forms.
          ROUND(AVG(form_duration_min) FILTER (WHERE form_duration_min IS NOT NULL AND form_duration_min > 0 AND form_duration_min < 360)::numeric, 2) AS avg_form_duration_min,
          (SELECT ROUND(percentile_cont(0.5)  WITHIN GROUP (ORDER BY form_duration_min)::numeric, 2)
             FROM mda_households {where}
              AND form_duration_min IS NOT NULL AND form_duration_min > 0 AND form_duration_min < 360) AS median_form_duration_min,
          (SELECT ROUND(percentile_cont(0.9)  WITHIN GROUP (ORDER BY form_duration_min)::numeric, 2)
             FROM mda_households {where}
              AND form_duration_min IS NOT NULL AND form_duration_min > 0 AND form_duration_min < 360) AS p90_form_duration_min,
          COUNT(*) FILTER (WHERE form_duration_min IS NOT NULL AND form_duration_min > 0 AND form_duration_min < 360) AS duration_sample_size
        FROM mda_households {where}
    """), params)
    row = result.fetchone()
    if row is None:
        return {
            "total_forms": 0, "duplicates": 0, "duplicate_gps": 0,
            "gps_outside_lga": 0, "gps_outside_ward": 0, "gps_outside_state": 0,
            "gps_poor_accuracy": 0, "gps_zero": 0, "after_hours": 0,
            "fast_forms": 0, "slow_forms": 0, "sync_lag": 0, "refusals": 0,
            "refusal_pct": 0.0, "forms_with_error": 0, "error_rate_pct": 0.0,
            "total_qc_flags": 0,
            "days_active": 0, "lgas_covered": 0,
            "ra_count": 0, "total_individuals": 0, "data_as_of": None,
            "avg_form_duration_min": None, "median_form_duration_min": None,
            "p90_form_duration_min": None, "duration_sample_size": 0,
        }
    keys = [
        "total_forms", "duplicates", "duplicate_gps",
        "gps_outside_lga", "gps_outside_ward", "gps_outside_state",
        "gps_poor_accuracy", "gps_zero", "after_hours",
        "fast_forms", "slow_forms", "sync_lag",
        "refusals", "refusal_pct", "forms_with_error",
        "days_active", "lgas_covered", "ra_count",
        "total_individuals", "data_as_of",
        "avg_form_duration_min", "median_form_duration_min",
        "p90_form_duration_min", "duration_sample_size",
    ]
    data = dict(zip(keys, row))
    # Ensure numeric types are JSON-safe
    int_keys = {"total_forms", "duplicates", "duplicate_gps", "gps_outside_lga",
                "gps_outside_ward", "gps_outside_state", "gps_poor_accuracy",
                "gps_zero", "after_hours", "fast_forms", "slow_forms", "sync_lag",
                "refusals", "forms_with_error", "days_active", "lgas_covered",
                "ra_count", "total_individuals", "duration_sample_size"}
    float_keys = {"refusal_pct", "avg_form_duration_min", "median_form_duration_min",
                  "p90_form_duration_min"}
    for k in keys:
        if k == "data_as_of":
            continue
        if k in int_keys:
            # Coalesce NULL → 0. SUM(CASE …) returns NULL when the row set is
            # empty (fresh project, before any forms have synced) — we don't
            # want that None cascading into arithmetic below or into the
            # dashboard's KPI tiles as "None".
            data[k] = int(data[k] or 0)
        elif k in float_keys and data[k] is not None:
            data[k] = float(data[k])
    if data["data_as_of"] is not None:
        data["data_as_of"] = str(data["data_as_of"])
    # Unified roll-ups so every page reads the same numbers.
    data["total_qc_flags"] = (
        data["fast_forms"] + data["slow_forms"] + data["after_hours"]
        + data["gps_outside_lga"] + data["gps_poor_accuracy"] + data["duplicate_gps"]
    )
    data["error_rate_pct"] = (
        round(100.0 * data["forms_with_error"] / data["total_forms"], 1)
        if data["total_forms"] else 0.0
    )
    return data


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/qc/ra-performance
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/qc/ra-performance")
async def qc_ra_performance(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    # date_trt (treatment date as entered by RA) replaced with received_on
    # so every "by date" surface in the dashboard uses the same column.
    params: dict = {"pid": pid}
    lga_sql = _lga_and(allowed_lgas_of(_u), "lga", params)
    result = await db.execute(text(f"""
        SELECT
          data_entry_persons,
          phone_number_data,
          lga,
          (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date::text AS date_trt,
          ra_key,
          COUNT(*) AS forms_submitted,
          SUM(CASE WHEN flag_refusal THEN 1 ELSE 0 END) AS refusals,
          ROUND(AVG(form_duration_min)::numeric, 1) AS avg_duration_min
        FROM mda_households
        WHERE project_id = :pid AND ra_key IS NOT NULL{lga_sql}
        GROUP BY data_entry_persons, phone_number_data, lga,
                 (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date, ra_key
        ORDER BY lga,
                 (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date, data_entry_persons
    """), params)
    rows = result.fetchall()
    keys = [
        "data_entry_persons", "phone_number_data", "lga", "date_trt",
        "ra_key", "forms_submitted", "refusals", "avg_duration_min",
    ]
    out = []
    for row in rows:
        d = dict(zip(keys, row))
        d["forms_submitted"] = int(d["forms_submitted"] or 0)
        d["refusals"] = int(d["refusals"] or 0)
        d["avg_duration_min"] = float(d["avg_duration_min"]) if d["avg_duration_min"] is not None else None
        d["meets_target"] = d["forms_submitted"] >= 10
        out.append(d)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/qc/refusals-by-lga
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/qc/refusals-by-lga")
async def qc_refusals_by_lga(
    lga: Optional[str] = None,
    ward: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    params: dict = {}
    filters: list = []
    if lga:  filters.append("lga = :lga");  params["lga"] = lga
    if ward: filters.append("REPLACE(LOWER(TRIM(ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')"); params["ward"] = ward
    where = _scoped_where(pid, filters, params, lgas=allowed_lgas_of(_u))
    result = await db.execute(text(f"""
        SELECT
          lga,
          COUNT(*) AS total_forms,
          SUM(CASE WHEN flag_refusal THEN 1 ELSE 0 END) AS refusals,
          ROUND(
            100.0 * SUM(CASE WHEN flag_refusal THEN 1 ELSE 0 END)
            / NULLIF(COUNT(*), 0), 1
          ) AS refusal_pct
        FROM mda_households {where}
        GROUP BY lga
        ORDER BY refusal_pct DESC
    """), params)
    rows = result.fetchall()
    return [
        {
            "lga": r[0],
            "total_forms": int(r[1] or 0),
            "refusals": int(r[2] or 0),
            "refusal_pct": float(r[3] or 0),
        }
        for r in rows
    ]


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/qc/duration-by-lga
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/qc/duration-by-lga")
async def qc_duration_by_lga(
    lga: Optional[str] = None,
    ward: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    params: dict = {}
    filters: list = ["form_duration_min IS NOT NULL"]
    if lga:  filters.append("lga = :lga");  params["lga"] = lga
    if ward: filters.append("REPLACE(LOWER(TRIM(ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')"); params["ward"] = ward
    where = _scoped_where(pid, filters, params, lgas=allowed_lgas_of(_u))
    result = await db.execute(text(f"""
        SELECT
          lga,
          COUNT(*) AS total,
          SUM(CASE WHEN flag_fast_form THEN 1 ELSE 0 END) AS fast_count,
          SUM(CASE WHEN flag_slow_form THEN 1 ELSE 0 END) AS slow_count,
          ROUND(AVG(form_duration_min)::numeric, 1) AS avg_min,
          ROUND(MIN(form_duration_min)::numeric, 1) AS min_min,
          ROUND(MAX(form_duration_min)::numeric, 1) AS max_min
        FROM mda_households {where}
        GROUP BY lga
        ORDER BY avg_min ASC
    """), params)
    rows = result.fetchall()
    return [
        {
            "lga": r[0],
            "total": int(r[1] or 0),
            "fast_count": int(r[2] or 0),
            "slow_count": int(r[3] or 0),
            "avg_min": float(r[4]) if r[4] is not None else None,
            "min_min": float(r[5]) if r[5] is not None else None,
            "max_min": float(r[6]) if r[6] is not None else None,
        }
        for r in rows
    ]


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/qc/duration-histogram  — distribution of form completion time
# Used by the bell-curve chart in Quality Checks to expose outliers visually.
# Buckets: 1-min wide from 0 → 60 min; any form ≥ 60 min lumped into the
# rightmost bin so the long tail doesn't squash the meaningful range.
# Returns bins + mean + stddev + p10/p90 so the client can overlay the
# "normal range" band on the histogram.
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/qc/duration-histogram")
async def qc_duration_histogram(
    lga: Optional[str] = None,
    ward: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    params: dict = {}
    filters: list = [
        "form_duration_min IS NOT NULL",
        "form_duration_min > 0",
        # Hard cap at 6 hours — anything beyond is almost certainly a device
        # left open / not a real form; including it would skew the mean and
        # collapse the visible bell.
        "form_duration_min < 360",
    ]
    if lga:  filters.append("lga = :lga");        params["lga"]  = lga
    if ward: filters.append("REPLACE(LOWER(TRIM(ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')"); params["ward"] = ward
    if date_from: filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date >= :date_from"); params["date_from"] = _as_date(date_from)
    if date_to:   filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date <= :date_to");   params["date_to"]   = _as_date(date_to)
    where = _scoped_where(pid, filters, params, lgas=allowed_lgas_of(_u))

    # 1-minute-wide bins from 0 → 60 min; >=60 lumped into bin 60 (the long tail).
    result = await db.execute(text(f"""
        SELECT
          LEAST(FLOOR(form_duration_min)::int, 60) AS bin_min,
          COUNT(*) AS n
        FROM mda_households {where}
        GROUP BY bin_min
        ORDER BY bin_min
    """), params)
    bins = [{"low": int(r[0]), "high": int(r[0]) + 1, "count": int(r[1])} for r in result.fetchall()]

    # Stats for the overlay (mean / median / stddev / percentiles).
    s_result = await db.execute(text(f"""
        SELECT
          ROUND(AVG(form_duration_min)::numeric, 2)               AS mean_min,
          ROUND(STDDEV_SAMP(form_duration_min)::numeric, 2)       AS stddev_min,
          ROUND(percentile_cont(0.5) WITHIN GROUP (ORDER BY form_duration_min)::numeric, 2) AS median_min,
          ROUND(percentile_cont(0.1) WITHIN GROUP (ORDER BY form_duration_min)::numeric, 2) AS p10_min,
          ROUND(percentile_cont(0.9) WITHIN GROUP (ORDER BY form_duration_min)::numeric, 2) AS p90_min,
          COUNT(*)                                                AS total
        FROM mda_households {where}
    """), params)
    s = s_result.fetchone()
    mean   = float(s[0]) if s and s[0] is not None else 0.0
    stddev = float(s[1]) if s and s[1] is not None else 0.0
    return {
        "bins": bins,
        "mean":   mean,
        "stddev": stddev,
        "median": float(s[2]) if s and s[2] is not None else 0.0,
        "p10":    float(s[3]) if s and s[3] is not None else 0.0,
        "p90":    float(s[4]) if s and s[4] is not None else 0.0,
        "lower_band": max(0.0, round(mean - 2 * stddev, 2)),
        "upper_band": round(mean + 2 * stddev, 2),
        "total":  int(s[5]) if s and s[5] is not None else 0,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/submissions/ward  — ward-level drill-down for charts
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/submissions/ward")
async def submissions_by_ward(
    lga:  Optional[str] = None,
    ward: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Ward-level submission & treatment counts, used for chart drill-down.

    Anchored on the canonical ward list in ``mda_baseline`` — the SAME source the
    ward *filter* dropdown uses (see ``/wards``) — NOT the free-text ``ward_name``
    typed on households. Household text is dirty: settlement / polling-unit names
    leak in (e.g. "Kwachiri" under Ungogo, which isn't a ward) and spellings drift
    ("Yada Kunya" vs the canonical "Yadakunya"), so grouping on it lists more
    "wards" than the LGA really has and splits real wards across spellings. We
    JOIN households to the baseline on a case- and space-insensitive key and
    report the canonical ward name, so variants fold into the real ward and
    non-wards drop out — the drill now matches the filter exactly.
    """
    filters = ["h.ward_name IS NOT NULL"]
    params: dict = {}
    if lga:  filters.append("h.lga = :lga"); params["lga"] = lga
    if ward:
        # Match the selected ward space/case-insensitively so canonical filter
        # values (e.g. "Yadakunya") still match dirty household text ("Yada Kunya").
        filters.append("REPLACE(LOWER(TRIM(h.ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')")
        params["ward"] = ward
    where = _scoped_where(pid, filters, params, alias="h", lgas=allowed_lgas_of(_u))
    result = await db.execute(text(f"""
        SELECT
          b.ward_disp                            AS ward_name,
          INITCAP(LOWER(TRIM(h.lga)))            AS lga,
          COUNT(*)                               AS forms,
          COALESCE(SUM(h.number_of_treated), 0)  AS treated,
          COUNT(DISTINCT h.hq_user)              AS teams,
          ROUND(COUNT(*)::numeric / NULLIF(COUNT(DISTINCT (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date), 0), 1) AS avg_per_day
        FROM mda_households h
        JOIN (
          -- De-duplicated to one row per (lga, ward) so the join can't fan out
          -- the household counts even if the baseline has settlement-level rows.
          SELECT
            REPLACE(LOWER(TRIM(lga)),  ' ', '') AS lga_key,
            REPLACE(LOWER(TRIM(ward)), ' ', '') AS ward_key,
            MIN(INITCAP(LOWER(TRIM(ward))))     AS ward_disp
          FROM mda_baseline
          WHERE project_id = :pid
            AND ward IS NOT NULL AND TRIM(ward) <> ''
            AND lga  IS NOT NULL AND TRIM(lga)  <> ''
          GROUP BY 1, 2
        ) b
          ON b.lga_key  = REPLACE(LOWER(TRIM(h.lga)),       ' ', '')
         AND b.ward_key = REPLACE(LOWER(TRIM(h.ward_name)), ' ', '')
        {where}
        GROUP BY b.ward_disp, INITCAP(LOWER(TRIM(h.lga)))
        ORDER BY forms DESC
    """), params)
    keys = ["ward_name", "lga", "forms", "treated", "teams", "avg_per_day"]
    return [dict(zip(keys, row)) for row in result.fetchall()]


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/qc/teams-summary  — per-team QC error breakdown
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/qc/teams-summary")
async def qc_teams_summary(
    lga:       Optional[str] = None,
    ward:      Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Per-team: total forms, forms with ≥1 error, error rate, and error type counts."""
    filters: list = ["hq_user IS NOT NULL"]
    params: dict = {}
    if lga:       filters.append("lga = :lga");       params["lga"]  = lga
    if ward:      filters.append("REPLACE(LOWER(TRIM(ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')"); params["ward"] = ward
    if date_from: filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date >= :date_from"); params["date_from"] = _as_date(date_from)
    if date_to:   filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date <= :date_to");   params["date_to"]   = _as_date(date_to)
    where = _scoped_where(pid, filters, params, lgas=allowed_lgas_of(_u))
    # forms_with_error uses the same definition as /qc/summary and /overview
    # (refusal NOT counted — it's an outcome; sync_lag NOT counted — it's a
    # connectivity issue, not data quality). sync_lag is still returned as a
    # separate column so the table can show it without inflating error_rate.
    result = await db.execute(text(f"""
        SELECT
          hq_user,
          lga,
          COUNT(*)                                                        AS total_forms,
          SUM(CASE WHEN (
            flag_duplicate OR flag_duplicate_gps OR flag_gps_outside_lga
            OR flag_gps_poor_accuracy OR flag_gps_zero OR flag_after_hours
            OR flag_fast_form OR flag_slow_form
          ) THEN 1 ELSE 0 END)                                           AS forms_with_error,
          SUM(CASE WHEN flag_duplicate         THEN 1 ELSE 0 END)        AS dup_forms,
          SUM(CASE WHEN flag_duplicate_gps     THEN 1 ELSE 0 END)        AS dup_gps,
          SUM(CASE WHEN flag_gps_outside_lga   THEN 1 ELSE 0 END)        AS gps_outside_lga,
          SUM(CASE WHEN flag_gps_poor_accuracy THEN 1 ELSE 0 END)        AS poor_gps,
          SUM(CASE WHEN flag_after_hours       THEN 1 ELSE 0 END)        AS after_hours,
          SUM(CASE WHEN flag_fast_form         THEN 1 ELSE 0 END)        AS fast_forms,
          SUM(CASE WHEN flag_slow_form         THEN 1 ELSE 0 END)        AS slow_forms,
          SUM(CASE WHEN flag_sync_lag          THEN 1 ELSE 0 END)        AS sync_lag,
          SUM(CASE WHEN flag_refusal           THEN 1 ELSE 0 END)        AS refusals,
          -- Per-team average form completion time. Same clipping window as
          -- /qc/summary and /qc/duration-histogram (>0 and <360 min) so a
          -- device-left-open outlier can't blow up one team's average.
          ROUND(AVG(form_duration_min) FILTER (
              WHERE form_duration_min IS NOT NULL
                AND form_duration_min > 0
                AND form_duration_min < 360
          )::numeric, 2)                                                 AS avg_duration_min
        FROM mda_households {where}
        GROUP BY hq_user, lga
        ORDER BY forms_with_error DESC, total_forms DESC
    """), params)
    rows = result.fetchall()
    keys = [
        "hq_user","lga","total_forms","forms_with_error",
        "dup_forms","dup_gps","gps_outside_lga","poor_gps",
        "after_hours","fast_forms","slow_forms","sync_lag","refusals",
        "avg_duration_min",
    ]
    out = []
    for row in rows:
        d = dict(zip(keys, row))
        # int-coerce the count columns (skip the float duration column)
        for k in keys[2:-1]:
            d[k] = int(d[k] or 0)
        d["avg_duration_min"] = float(d["avg_duration_min"]) if d["avg_duration_min"] is not None else None
        d["error_rate"] = round(100.0 * d["forms_with_error"] / d["total_forms"], 1) if d["total_forms"] else 0
        out.append(d)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/qc/gps/geojson
# ─────────────────────────────────────────────────────────────────────────────

async def _gps_geojson(db: AsyncSession, pid: int, where_clause: str, limit: int = 20000, lgas=None):
    """Shared helper for filtered GPS GeoJSON endpoints. ``lgas`` (from
    allowed_lgas_of) restricts the points to an LGA-scoped account's LGA(s)."""
    # Map tooltip shows received_on (the platform's date of record). The
    # outer property key stays `date_trt` so the frontend tooltip template
    # doesn't have to change.
    params: dict = {"pid": pid}
    lga_sql = _lga_and(lgas, "lga", params)
    result = await db.execute(text(f"""
        SELECT
          formid, lga, data_entry_persons,
          (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date::text AS date_trt,
          gps_accuracy, form_duration_min,
          flag_gps_outside_lga, flag_gps_outside_ward,
          flag_gps_outside_state, flag_duplicate_gps,
          latitude, longitude,
          ST_AsGeoJSON(geom)::json AS geometry
        FROM mda_households
        WHERE project_id = :pid AND geom IS NOT NULL AND {where_clause}{lga_sql}
        LIMIT {limit}
    """), params)
    rows = result.fetchall()
    features = []
    for row in rows:
        (formid, lga, ra, date_trt, acc, dur,
         out_lga, out_ward, out_state, dup_gps, lat, lon, geometry) = row
        features.append({
            "type": "Feature",
            "geometry": geometry,
            "properties": {
                "formid": formid, "lga": lga, "ra": ra,
                "date_trt": date_trt,
                "accuracy": float(acc) if acc else None,
                "duration_min": float(dur) if dur else None,
                "out_lga": bool(out_lga), "out_ward": bool(out_ward),
                "out_state": bool(out_state), "dup_gps": bool(dup_gps),
                "lat": float(lat) if lat else None,
                "lon": float(lon) if lon else None,
            },
        })
    return {"type": "FeatureCollection", "features": features}


@router.get("/qc/gps/outside-lga")
async def gps_outside_lga(pid: int = Depends(resolve_pid), db: AsyncSession = Depends(get_db), _u: Optional[User] = Depends(get_current_user_optional)):
    return await _gps_geojson(db, pid, "flag_gps_outside_lga = TRUE", lgas=allowed_lgas_of(_u))


@router.get("/qc/gps/outside-ward")
async def gps_outside_ward(pid: int = Depends(resolve_pid), db: AsyncSession = Depends(get_db), _u: Optional[User] = Depends(get_current_user_optional)):
    return await _gps_geojson(db, pid, "flag_gps_outside_ward = TRUE", lgas=allowed_lgas_of(_u))


@router.get("/qc/gps/outside-state")
async def gps_outside_state(pid: int = Depends(resolve_pid), db: AsyncSession = Depends(get_db), _u: Optional[User] = Depends(get_current_user_optional)):
    return await _gps_geojson(db, pid, "flag_gps_outside_state = TRUE", lgas=allowed_lgas_of(_u))


@router.get("/qc/gps/duplicate")
async def gps_duplicate(pid: int = Depends(resolve_pid), db: AsyncSession = Depends(get_db), _u: Optional[User] = Depends(get_current_user_optional)):
    return await _gps_geojson(db, pid, "flag_duplicate_gps = TRUE", lgas=allowed_lgas_of(_u))


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/overview
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/overview")
async def mda_overview(
    lga:       Optional[str] = None,
    ward:      Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    filters, params = [], {}
    if lga:       filters.append("lga = :lga");       params["lga"]  = lga
    if ward:      filters.append("REPLACE(LOWER(TRIM(ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')"); params["ward"] = ward
    if date_from: filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date >= :date_from"); params["date_from"] = _as_date(date_from)
    if date_to:   filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date <= :date_to");   params["date_to"]   = _as_date(date_to)
    lgas = allowed_lgas_of(_u)
    where = _scoped_where(pid, filters, params, lgas=lgas)
    lga_bl = _lga_and(lgas, "lga", params)  # for the mda_baseline sub-query below
    result = await db.execute(text(f"""
        SELECT
          COUNT(*) AS total_forms,
          COALESCE(SUM(number_of_treated), 0) AS total_treated,
          COUNT(DISTINCT hq_user) AS teams_active,
          COUNT(DISTINCT (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date) AS days_active,
          COUNT(DISTINCT lga) AS lgas_covered,
          SUM(CASE WHEN flag_refusal           THEN 1 ELSE 0 END) AS refusals,
          SUM(CASE WHEN flag_fast_form         THEN 1 ELSE 0 END) AS fast_forms,
          SUM(CASE WHEN flag_slow_form         THEN 1 ELSE 0 END) AS slow_forms,
          SUM(CASE WHEN flag_after_hours       THEN 1 ELSE 0 END) AS after_hours,
          SUM(CASE WHEN flag_gps_outside_lga   THEN 1 ELSE 0 END) AS gps_outside_lga,
          SUM(CASE WHEN flag_gps_poor_accuracy THEN 1 ELSE 0 END) AS gps_poor_accuracy,
          SUM(CASE WHEN flag_duplicate_gps     THEN 1 ELSE 0 END) AS duplicate_gps,
          -- forms_with_any_error: at least one *real* error flag (refusal
          -- excluded — it is an outcome, not a data-quality issue).
          SUM(CASE WHEN (
              flag_fast_form OR flag_slow_form OR flag_after_hours
              OR flag_gps_outside_lga OR flag_gps_poor_accuracy
              OR flag_duplicate_gps OR flag_duplicate OR flag_gps_zero
          ) THEN 1 ELSE 0 END) AS forms_with_error,
          (SELECT COALESCE(SUM(total_treated),0) FROM mda_baseline WHERE project_id = :pid{lga_bl}) AS baseline_total,
          MIN((received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)::text AS campaign_start,
          MAX((received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)::text AS campaign_end
        FROM mda_households {where}
    """), params)
    row = result.fetchone()
    if not row or not row[0]:
        return {"total_forms": 0}
    d = dict(zip([
        "total_forms","total_treated","teams_active","days_active",
        "lgas_covered","refusals","fast_forms","slow_forms","after_hours",
        "gps_outside_lga","gps_poor_accuracy","duplicate_gps",
        "forms_with_error","baseline_total","campaign_start","campaign_end"
    ], row))
    for k in ["total_forms","total_treated","teams_active","days_active","lgas_covered",
              "refusals","fast_forms","slow_forms","after_hours",
              "gps_outside_lga","gps_poor_accuracy","duplicate_gps",
              "forms_with_error","baseline_total"]:
        if d[k] is not None: d[k] = int(d[k])
    bl = d["baseline_total"] or 0
    d["coverage_pct"] = round(100.0 * d["total_treated"] / bl, 1) if bl > 0 else 0
    # Unified definition: sum of *real* error-flag instances. A form can
    # contribute to multiple flags so this is the total number of QC issues,
    # not the number of flagged forms (that is forms_with_error). Refusals
    # are excluded — they are tracked separately as a campaign outcome.
    d["total_qc_flags"] = (
        d["fast_forms"] + d["slow_forms"] + d["after_hours"]
        + d["gps_outside_lga"] + d["gps_poor_accuracy"] + d["duplicate_gps"]
    )
    d["error_rate_pct"] = round(100.0 * d["forms_with_error"] / d["total_forms"], 1) if d["total_forms"] else 0
    # Planned campaign window — pulled from geo_projects.campaign_*_date.
    # current_campaign_day is computed from today's date in Africa/Lagos,
    # NOT from the count of submission-days. That way the Day-N tile ticks
    # forward every calendar day even before the first form arrives, and
    # stays clamped to [1, planned_duration_days] (or open-ended N if no
    # end date is set). The legacy `days_active` field (count of distinct
    # submission days) is preserved for callers that still want it.
    proj_res = await db.execute(text("""
        SELECT campaign_start_date, campaign_end_date,
               (NOW() AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date AS today_lagos
        FROM geo_projects WHERE id = :pid
    """), {"pid": pid})
    proj_row = proj_res.fetchone()
    if proj_row:
        d["planned_start_date"] = proj_row[0].isoformat() if proj_row[0] else None
        d["planned_end_date"]   = proj_row[1].isoformat() if proj_row[1] else None
        if proj_row[0] and proj_row[1]:
            d["planned_duration_days"] = (proj_row[1] - proj_row[0]).days + 1
        else:
            d["planned_duration_days"] = None
        if proj_row[0]:
            today = proj_row[2]
            raw_day = (today - proj_row[0]).days + 1
            if raw_day < 1:
                d["current_campaign_day"] = 0  # campaign hasn't started yet
            elif d.get("planned_duration_days"):
                d["current_campaign_day"] = min(raw_day, d["planned_duration_days"])
            else:
                d["current_campaign_day"] = raw_day
        else:
            d["current_campaign_day"] = None

    # Forms-target estimator. The dashboard's "Forms Submitted vs Target"
    # column chart needs a campaign-wide target form count. We derive it
    # from the SARMAAN target rate of 80 forms / team / day. Fallback to
    # the boundary settlement count (each settlement should be visited at
    # least once) so projects without a campaign-window configured still
    # get a meaningful target.
    # Forms target = the delivery-capacity estimate ONLY: active teams × 80
    # forms/team/day × planned campaign-days. If there's no campaign window
    # (no planned_duration_days) or no teams reporting yet, we return None so
    # the card shows "no target yet" rather than a misleading value. We used
    # to fall back to the boundary settlement count, but that isn't an
    # accurate forms target - a value must be accurate or absent.
    target_forms: Optional[int] = None
    teams = d.get("teams_active") or 0
    pdd = d.get("planned_duration_days")
    if teams > 0 and pdd:
        target_forms = teams * 80 * int(pdd)
    d["target_forms"] = target_forms

    # Repeat-group drift QC. Counts households whose form-level
    # number_of_treated disagrees with the count of treated child rows in
    # the repeat group. A non-zero value points to under-filled forms /
    # missing repeat-group rows / status mis-entries — actionable for the
    # field-supervisor team.
    drift_res = await db.execute(text(f"""
        WITH ind AS (
          SELECT i.hh_formid,
                 COUNT(*) FILTER (WHERE i.treatment_status = '1') AS treated_rows
          FROM mda_individuals i
          WHERE i.project_id = :pid
          GROUP BY i.hh_formid
        )
        SELECT
          COUNT(*) FILTER (WHERE
              COALESCE(h.number_of_treated, 0) <> COALESCE(ind.treated_rows, 0)
          ) AS drift_households,
          COUNT(*) AS total_households,
          COALESCE(SUM(GREATEST(
              COALESCE(h.number_of_treated, 0) - COALESCE(ind.treated_rows, 0), 0
          )), 0) AS missing_repeat_rows
        FROM mda_households h
        LEFT JOIN ind ON ind.hh_formid = h.formid
        {where}
    """), params)
    drift_row = drift_res.fetchone()
    if drift_row:
        d["drift_households"]     = int(drift_row[0] or 0)
        d["drift_missing_rows"]   = int(drift_row[2] or 0)
        d["drift_pct"] = round(100.0 * d["drift_households"] / d["total_forms"], 1) if d["total_forms"] else 0.0
    else:
        d["drift_households"] = 0
        d["drift_missing_rows"] = 0
        d["drift_pct"] = 0.0
    return d


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/trends/daily
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/trends/daily")
async def mda_trends_daily(
    lga:       Optional[str] = None,
    ward:      Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    filters = ["received_on IS NOT NULL"]
    params: dict = {}
    if lga:       filters.append("lga = :lga");       params["lga"]  = lga
    if ward:      filters.append("REPLACE(LOWER(TRIM(ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')"); params["ward"] = ward
    if date_from: filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date >= :date_from"); params["date_from"] = _as_date(date_from)
    if date_to:   filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date <= :date_to");   params["date_to"]   = _as_date(date_to)
    where = _scoped_where(pid, filters, params, lgas=allowed_lgas_of(_u))
    result = await db.execute(text(f"""
        SELECT
          (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date::text AS date,
          COUNT(*) AS forms,
          COALESCE(SUM(number_of_treated), 0) AS treated,
          COUNT(DISTINCT hq_user) AS teams
        FROM mda_households {where}
        GROUP BY (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
        ORDER BY (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
    """), params)
    return [dict(zip(["date","forms","treated","teams"], row)) for row in result.fetchall()]


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/teams/performance
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/teams/performance")
async def mda_teams(
    lga:  Optional[str] = None,
    ward: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    filters = ["hq_user IS NOT NULL"]
    params: dict = {}
    if lga:  filters.append("lga = :lga");       params["lga"]  = lga
    if ward: filters.append("REPLACE(LOWER(TRIM(ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')"); params["ward"] = ward
    where = _scoped_where(pid, filters, params, lgas=allowed_lgas_of(_u))
    result = await db.execute(text(f"""
        SELECT
          hq_user,
          lga,
          COUNT(*) AS total_forms,
          COUNT(DISTINCT (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date) AS days_active,
          ROUND(COUNT(*)::numeric / NULLIF(COUNT(DISTINCT (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date), 0), 1) AS avg_per_day,
          COALESCE(SUM(number_of_treated), 0) AS total_treated,
          SUM(CASE WHEN flag_after_hours THEN 1 ELSE 0 END) AS after_hours,
          SUM(CASE WHEN flag_fast_form THEN 1 ELSE 0 END) AS fast_forms
        FROM mda_households {where}
        GROUP BY hq_user, lga
        ORDER BY avg_per_day DESC
    """), params)
    rows = result.fetchall()
    keys = ["hq_user","lga","total_forms","days_active","avg_per_day","total_treated","after_hours","fast_forms"]
    out = []
    for row in rows:
        d = dict(zip(keys, row))
        d["avg_per_day"] = float(d["avg_per_day"] or 0)
        d["meets_target"] = d["avg_per_day"] >= 80
        for k in ["total_forms","days_active","total_treated","after_hours","fast_forms"]:
            if d[k] is not None: d[k] = int(d[k])
        out.append(d)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/coverage/lga
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/coverage/lga")
async def mda_coverage_lga(
    lga:       Optional[str] = None,
    ward:      Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Coverage per LGA. The universe of LGAs is the union of:
      - baseline-targeted LGAs
      - LGAs that have submitted at least one form
      - the project's in-scope LGA list (see app.services.project_scope) —
        pulled from the boundary table so a data-collection LGA appears
        even before any target is uploaded or any form is synced.

    Why a union of all three: pre-rollout (no baseline yet) a state still
    wants the LGA bar chart to show the forms it's already syncing. Post-
    baseline a targeted LGA must still appear even if zero forms have
    arrived. And a scope-registered LGA (Kano R3's 36) must appear even if
    it has neither a baseline row nor any household forms yet — otherwise
    the operator's "36 collection LGAs" get collapsed to whatever subset
    the baseline sheet happened to cover.
    """
    from app.services.project_scope import in_scope_lgas_for  # local — avoids circular
    from app.services.spatial_engine import _resolve_boundary_pid

    # Filters that apply on the household side (date/ward). pid is added
    # separately because it gates the household subquery directly.
    hh_filters: list = ["h.project_id = :pid"]
    params: dict = {"pid": pid}
    if ward:
        hh_filters.append("REPLACE(LOWER(TRIM(h.ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')")
        params["ward"] = ward
    if date_from:
        hh_filters.append("(h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date >= :date_from")
        params["date_from"] = _as_date(date_from)
    if date_to:
        hh_filters.append("(h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date <= :date_to")
        params["date_to"] = _as_date(date_to)
    hh_where = " AND ".join(hh_filters)

    # Optional LGA-name filter applies to both baseline and households.
    lga_filter_baseline = ""
    lga_filter_hh = ""
    if lga:
        lga_filter_baseline = "AND UPPER(TRIM(lga)) = UPPER(TRIM(:lga))"
        lga_filter_hh = "AND UPPER(TRIM(h.lga)) = UPPER(TRIM(:lga))"
        params["lga"] = lga
    # LGA-scoped account: further restrict the LGA universe to its own LGA(s).
    lgas = allowed_lgas_of(_u)
    lga_filter_baseline += _lga_and(lgas, "lga", params)
    lga_filter_hh += _lga_and(lgas, "h.lga", params)

    # Look up whether this project has a registered "in-scope LGA" list
    # (currently only Kano R3). If yes, we'll pull that LGA set from the
    # boundary table as a third contributor to the universe CTE below.
    scope_names_lower: Optional[list[str]] = None
    p_row = (await db.execute(
        text("SELECT state_name, round_number FROM geo_projects WHERE id = :pid"),
        {"pid": pid},
    )).fetchone()
    if p_row:
        scope_set = in_scope_lgas_for(p_row[0], p_row[1])
        if scope_set:
            scope_names_lower = [n.strip().lower() for n in scope_set]

    # Conditionally add a `scope_lgas` CTE + UNION arm. Kept out of the SQL
    # entirely when scope_names_lower is None so Sokoto R4/R5/Kano Pilot see
    # exactly the same query they've always seen.
    scope_cte      = ""
    scope_union    = ""
    if scope_names_lower is not None:
        boundary_pid = await _resolve_boundary_pid(pid, db)
        params["boundary_pid"]     = boundary_pid
        params["scope_names_lower"] = scope_names_lower
        # LGA-scoped user restriction (allowed_lgas) also applied here so an
        # LGA-scoped account never widens its view via the scope path.
        scope_lga_and = _lga_and(lgas, "l.lga_name", params)
        scope_cte = f"""
          scope_lgas AS (
            SELECT
              INITCAP(LOWER(TRIM(l.lga_name))) AS lga,
              UPPER(TRIM(l.lga_name))          AS lga_key
            FROM lgas l
            WHERE l.project_id = :boundary_pid
              AND LOWER(TRIM(l.lga_name)) = ANY(:scope_names_lower)
              {scope_lga_and}
          ),
        """
        scope_union = "UNION SELECT lga_key, lga FROM scope_lgas"

    result = await db.execute(text(f"""
        WITH baseline AS (
          SELECT
            INITCAP(LOWER(TRIM(lga))) AS lga,
            UPPER(TRIM(lga))          AS lga_key,
            SUM(total_treated)        AS baseline_total
          FROM mda_baseline
          WHERE project_id = :pid {lga_filter_baseline}
            AND lga IS NOT NULL AND TRIM(lga) <> ''
          GROUP BY 1, 2
        ),
        hh AS (
          SELECT
            INITCAP(LOWER(TRIM(h.lga))) AS lga,
            UPPER(TRIM(h.lga))          AS lga_key,
            COUNT(h.id)                 AS forms,
            COALESCE(SUM(h.number_of_treated), 0) AS actual_treated,
            COUNT(DISTINCT h.hq_user)   AS teams,
            COUNT(DISTINCT (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date) AS days_reported
          FROM mda_households h
          WHERE {hh_where} {lga_filter_hh}
            AND h.lga IS NOT NULL AND TRIM(h.lga) <> ''
          GROUP BY 1, 2
        ),
        {scope_cte}
        universe AS (
          SELECT lga_key, lga FROM baseline
          UNION
          SELECT lga_key, lga FROM hh
          {scope_union}
        )
        SELECT
          u.lga,
          COALESCE(hh.forms, 0)                  AS forms,
          COALESCE(hh.actual_treated, 0)         AS actual_treated,
          COALESCE(b.baseline_total, 0)          AS baseline_total,
          CASE WHEN COALESCE(b.baseline_total, 0) > 0
               THEN ROUND(100.0 * COALESCE(hh.actual_treated, 0) / b.baseline_total, 1)
               ELSE 0 END                        AS coverage_pct,
          COALESCE(hh.teams, 0)                  AS teams,
          COALESCE(hh.days_reported, 0)          AS days_reported
        FROM universe u
        LEFT JOIN baseline b  ON b.lga_key  = u.lga_key
        LEFT JOIN hh          ON hh.lga_key = u.lga_key
        ORDER BY coverage_pct DESC, u.lga
    """), params)
    rows = result.fetchall()
    keys = ["lga","forms","actual_treated","baseline_total","coverage_pct","teams","days_reported"]
    return [dict(zip(keys, row)) for row in rows]


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/coverage/ward
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/coverage/ward")
async def mda_coverage_ward(
    lga: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Ward-level administrative coverage, optionally filtered by LGA.

    Anchored on the ward-level TARGETS in mda_baseline (the authoritative ward
    list from the target workbook), not on the free-text ``ward_name`` typed on
    each household. Household ward text is dirty (settlement / polling-unit names
    leak in, and names repeat across LGAs), which used to surface ~40 "wards"
    for an LGA that really has ~11. Anchoring on the targets returns exactly the
    real wards, and coverage is the ward's own administrative coverage:
    treated-in-ward / ward-target.
    """
    params: dict = {"pid": pid}
    lgas = allowed_lgas_of(_u)
    b_lga = a_lga = ""
    if lga:
        b_lga = " AND UPPER(TRIM(mb.lga)) = UPPER(TRIM(:lga))"
        a_lga = " AND UPPER(TRIM(h.lga)) = UPPER(TRIM(:lga))"
        params["lga"] = lga
    b_scope = _lga_and(lgas, "mb.lga", params)   # per-user LGA access scope
    a_scope = _lga_and(lgas, "h.lga", params)
    result = await db.execute(text(f"""
        WITH tgt AS (
          -- Space-insensitive keys so canonical baseline wards match dirty
          -- household spellings ("Yadakunya" vs "Yada Kunya"); dedup to one row
          -- per (lga, ward) with MIN() display names + summed target.
          SELECT REPLACE(UPPER(TRIM(mb.lga)),  ' ', '') AS lga_key,
                 MIN(INITCAP(LOWER(TRIM(mb.lga))))       AS lga,
                 REPLACE(UPPER(TRIM(mb.ward)), ' ', '')  AS ward_key,
                 MIN(INITCAP(LOWER(TRIM(mb.ward))))      AS ward_name,
                 SUM(mb.total_treated)                   AS baseline_total
          FROM mda_baseline mb
          WHERE mb.project_id = :pid AND mb.ward IS NOT NULL AND TRIM(mb.ward) <> ''{b_lga}{b_scope}
          GROUP BY 1, 3
        ),
        act AS (
          SELECT REPLACE(UPPER(TRIM(h.lga)),       ' ', '') AS lga_key,
                 REPLACE(UPPER(TRIM(h.ward_name)), ' ', '') AS ward_key,
                 COUNT(*)                 AS forms,
                 COALESCE(SUM(h.number_of_treated), 0) AS actual_treated,
                 COUNT(DISTINCT h.hq_user) AS teams,
                 COUNT(DISTINCT (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date) AS days_reported
          FROM mda_households h
          WHERE h.project_id = :pid AND h.ward_name IS NOT NULL{a_lga}{a_scope}
          GROUP BY 1, 2
        )
        SELECT t.ward_name, t.lga,
               COALESCE(a.forms, 0)          AS forms,
               COALESCE(a.actual_treated, 0) AS actual_treated,
               t.baseline_total,
               CASE WHEN t.baseline_total > 0
                    THEN ROUND(100.0 * COALESCE(a.actual_treated, 0) / t.baseline_total, 1)
                    ELSE 0 END               AS coverage_pct,
               COALESCE(a.teams, 0)          AS teams,
               COALESCE(a.days_reported, 0)  AS days_reported
        FROM tgt t
        LEFT JOIN act a ON a.lga_key = t.lga_key AND a.ward_key = t.ward_key
        ORDER BY coverage_pct DESC NULLS LAST
    """), params)
    rows = result.fetchall()
    keys = ["ward_name","lga","forms","actual_treated","baseline_total","coverage_pct","teams","days_reported"]
    return [dict(zip(keys, row)) for row in rows]


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/individuals/age-summary
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/individuals/age-summary")
async def individuals_age_summary(
    lga:       Optional[str] = None,
    ward:      Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Age-band breakdown from mda_individuals, filterable by LGA, ward and date.
    The date filter is honoured on every figure here (the reported treated total
    and the age bands) so the chart stays in step with the Overview KPIs when the
    operator picks a specific day/range."""
    hh_filters = ["h.project_id = :pid", "i.project_id = :pid", "h.lga IS NOT NULL"]
    params: dict = {"pid": pid}
    if lga:
        hh_filters.append("h.lga = :lga")
        params["lga"] = lga
    if ward:
        hh_filters.append("REPLACE(LOWER(TRIM(h.ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')")
        params["ward"] = ward
    if date_from:
        hh_filters.append("(h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date >= :date_from")
        params["date_from"] = _as_date(date_from)
    if date_to:
        hh_filters.append("(h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date <= :date_to")
        params["date_to"] = _as_date(date_to)
    lgas = allowed_lgas_of(_u)
    hh_where = " AND ".join(hh_filters) + _lga_and(lgas, "h.lga", params)
    lga_bl = _lga_and(lgas, "lga", params)  # for the mda_baseline sub-query below

    result = await db.execute(text(f"""
        SELECT
            COUNT(*) FILTER (WHERE i.age_in_months BETWEEN 1 AND 11)                          AS total_1_11,
            COUNT(*) FILTER (WHERE i.age_in_months BETWEEN 1 AND 11  AND i.treatment_status='1') AS treated_1_11,
            COUNT(*) FILTER (WHERE i.age_in_months BETWEEN 12 AND 59)                         AS total_12_59,
            COUNT(*) FILTER (WHERE i.age_in_months BETWEEN 12 AND 59 AND i.treatment_status='1') AS treated_12_59,
            COUNT(*) FILTER (WHERE i.treatment_status='1')                                     AS total_treated,
            COUNT(*)                                                                            AS grand_total
        FROM mda_individuals i
        JOIN mda_households h ON h.formid = i.hh_formid
        WHERE i.age_in_months IS NOT NULL AND i.age_in_months BETWEEN 1 AND 59
          AND {hh_where}
    """), params)
    row = result.fetchone()
    # Baselines: grand total + age-band breakdown (R5+ populates the breakdown columns).
    bl = await db.execute(text(f"""
        SELECT
          COALESCE(SUM(total_treated), 0)                                          AS total,
          COALESCE(SUM(COALESCE(target_1_11_f, 0) + COALESCE(target_1_11_m, 0)), 0)   AS baseline_1_11,
          COALESCE(SUM(COALESCE(target_12_59_f, 0) + COALESCE(target_12_59_m, 0)), 0) AS baseline_12_59
        FROM mda_baseline
        WHERE project_id = :pid{lga_bl}
    """), params)
    bl_row = bl.fetchone()

    # Reported treated total = SUM(number_of_treated) over the SAME households (no
    # individuals join, so no fan-out). This is the authoritative "children
    # treated" figure — the primary reported field and the coverage numerator —
    # so the 1–59 chart total matches the "Children Treated" KPI and Treatment
    # Coverage %. Scoped via _scoped_where (project + campaign-start-date bound +
    # LGA access), exactly like the Overview KPI, so the two agree to the unit.
    # It runs ~drift higher than the per-child row count above because some
    # treated children have no age-recorded repeat-group row (see the
    # Form ↔ Repeat-Group Drift QC), which is why the age *bands* (from
    # mda_individuals) can't sum exactly to it.
    rep_filters = []
    if lga:
        rep_filters.append("lga = :lga")
    if ward:
        rep_filters.append("REPLACE(LOWER(TRIM(ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')")
    if date_from:
        rep_filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date >= :date_from")
    if date_to:
        rep_filters.append("(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date <= :date_to")
    rep_where = _scoped_where(pid, rep_filters, params, lgas=lgas)
    rep = await db.execute(text(f"""
        SELECT COALESCE(SUM(number_of_treated), 0) AS treated_reported
        FROM mda_households
        {rep_where}
    """), params)
    treated_reported = int(rep.scalar() or 0)

    treated_1_11   = int(row.treated_1_11 or 0)
    treated_12_59  = int(row.treated_12_59 or 0)
    baseline_1_11  = int(bl_row.baseline_1_11) if bl_row else 0
    baseline_12_59 = int(bl_row.baseline_12_59) if bl_row else 0

    return {
        "total_1_11":    int(row.total_1_11 or 0),
        "treated_1_11":  treated_1_11,
        "total_12_59":   int(row.total_12_59 or 0),
        "treated_12_59": treated_12_59,
        "total_treated": int(row.total_treated or 0),
        "treated_reported": treated_reported,
        "grand_total":   int(row.grand_total or 0),
        "baseline_total": int(bl_row.total) if bl_row else 0,
        # Age-band coverage (populated only if the round's baseline carries per-band targets)
        "baseline_1_11":      baseline_1_11,
        "baseline_12_59":     baseline_12_59,
        "coverage_1_11_pct":  round(100.0 * treated_1_11  / baseline_1_11,  1) if baseline_1_11  > 0 else None,
        "coverage_12_59_pct": round(100.0 * treated_12_59 / baseline_12_59, 1) if baseline_12_59 > 0 else None,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/coverage/refusals-analysis
# Missed-household / non-compliance analysis. Returns the active round's
# refusal stats, reason-code breakdown, free-text reasons, per-LGA detail,
# and a R4-vs-R5 (or R(n) vs R(n-1)) round-over-round delta on refusal rate.
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/coverage/refusals-analysis")
async def coverage_refusals_analysis(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    # Round-over-round refusal rates. We also estimate missed children per
    # refusal: refusing households never enter the individual repeat-group so
    # mda_individuals has no rows for them. Instead we apply the round's
    # average children-per-consenting-household ratio as the best-available
    # estimator. The DB field is labeled `estimated_missed_individuals` to
    # keep the assumption explicit.
    # Per-round figures are bounded by each project's campaign_start_date when
    # set, so pre-campaign test rows don't inflate refusal totals or pull the
    # mean-children-per-household ratio off.
    # Scope rounds to the same state as the selected project — round-over-round
    # comparisons only make sense within a state (Kano Pilot is not "the round
    # before Sokoto R4"). Each row carries `is_selected` so the frontend can
    # pick the active project's summary deterministically rather than guessing.
    lgas = allowed_lgas_of(_u)
    params_r: dict = {"pid": pid}
    lga_h = _lga_and(lgas, "h.lga", params_r)  # each round sub-query is on mda_households h
    rounds_res = await db.execute(text(f"""
        SELECT
          p.id, p.round_number,
          (p.id = :pid) AS is_selected,
          (SELECT COUNT(*) FROM mda_households h
             WHERE h.project_id = p.id{lga_h}
               AND (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
                   >= COALESCE(p.campaign_start_date, '1900-01-01'::date)) AS total_forms,
          (SELECT COUNT(*) FROM mda_households h
             WHERE h.project_id = p.id AND h.flag_refusal{lga_h}
               AND (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
                   >= COALESCE(p.campaign_start_date, '1900-01-01'::date)) AS refusals,
          (SELECT COUNT(*) FROM mda_individuals i
              JOIN mda_households h ON h.formid = i.hh_formid AND h.project_id = i.project_id
              WHERE h.project_id = p.id AND NOT h.flag_refusal{lga_h}
                AND (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
                    >= COALESCE(p.campaign_start_date, '1900-01-01'::date)) AS consenting_individuals,
          (SELECT COUNT(*) FROM mda_households h
             WHERE h.project_id = p.id AND NOT h.flag_refusal{lga_h}
               AND (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
                   >= COALESCE(p.campaign_start_date, '1900-01-01'::date)) AS consenting_households
        FROM geo_projects p
        WHERE p.state_name = (SELECT state_name FROM geo_projects WHERE id = :pid)
        ORDER BY p.round_number NULLS LAST, p.id
    """), params_r)
    rounds_rows = []
    for r in rounds_res.fetchall():
        tf = int(r.total_forms or 0)
        rf = int(r.refusals or 0)
        consenting_ind = int(r.consenting_individuals or 0)
        consenting_hh  = int(r.consenting_households or 0)
        avg_children_per_hh = (consenting_ind / consenting_hh) if consenting_hh > 0 else 0
        rounds_rows.append({
            "project_id":   r.id,
            "is_selected":  bool(r.is_selected),
            "round_number": r.round_number,
            "round_label":  f"Round {r.round_number}" if r.round_number is not None else f"P#{r.id}",
            "total_forms":  tf,
            "refusals":     rf,
            "refusal_pct":  round(100.0 * rf / tf, 2) if tf > 0 else 0.0,
            "avg_children_per_household": round(avg_children_per_hh, 2),
            "estimated_missed_individuals": int(round(rf * avg_children_per_hh)),
        })

    # Per-reason breakdown — active round only
    params_re: dict = {"pid": pid}
    lga_l_re = _lga_and(lgas, "lga", params_re)
    reason_res = await db.execute(text(f"""
        SELECT reasons_for_refusal AS code, COUNT(*) AS cnt
        FROM mda_households
        WHERE project_id = :pid AND flag_refusal = TRUE{lga_l_re}
        GROUP BY reasons_for_refusal
        ORDER BY cnt DESC
    """), params_re)
    by_reason = [{"code": (r.code or 'unspecified'), "count": int(r.cnt)} for r in reason_res.fetchall()]

    # Free-text "other" reasons — case-folded so 'INSECURITY ISSUE' and
    # 'insecurity issue' aren't double-counted.
    params_fr: dict = {"pid": pid}
    lga_l_fr = _lga_and(lgas, "lga", params_fr)
    free_res = await db.execute(text(f"""
        SELECT LOWER(TRIM(others_reasons_for_refusal)) AS reason, COUNT(*) AS cnt
        FROM mda_households
        WHERE project_id = :pid AND flag_refusal = TRUE{lga_l_fr}
          AND others_reasons_for_refusal IS NOT NULL
          AND TRIM(others_reasons_for_refusal) <> ''
        GROUP BY LOWER(TRIM(others_reasons_for_refusal))
        ORDER BY cnt DESC LIMIT 15
    """), params_fr)
    free_text = [{"text": r.reason, "count": int(r.cnt)} for r in free_res.fetchall()]

    # Per-LGA refusal totals + missed children — active round only
    params_lg: dict = {"pid": pid}
    lga_l_lg = _lga_and(lgas, "lga", params_lg)
    lga_h_lg = _lga_and(lgas, "h.lga", params_lg)
    lga_res = await db.execute(text(f"""
        WITH r AS (
          SELECT INITCAP(TRIM(lga)) AS lga,
                 COUNT(*) AS total_forms,
                 COUNT(*) FILTER (WHERE flag_refusal) AS refusals
          FROM mda_households
          WHERE project_id = :pid AND lga IS NOT NULL{lga_l_lg}
          GROUP BY INITCAP(TRIM(lga))
        ),
        m AS (
          SELECT INITCAP(TRIM(h.lga)) AS lga, COUNT(*) AS missed
          FROM mda_individuals i
          JOIN mda_households h ON h.formid = i.hh_formid AND h.project_id = i.project_id
          WHERE h.project_id = :pid AND h.flag_refusal AND h.lga IS NOT NULL{lga_h_lg}
          GROUP BY INITCAP(TRIM(h.lga))
        )
        SELECT r.lga, r.total_forms, r.refusals,
               COALESCE(m.missed, 0) AS missed_individuals,
               CASE WHEN r.total_forms > 0
                    THEN ROUND(100.0 * r.refusals / r.total_forms, 2)
                    ELSE 0 END AS refusal_pct
        FROM r LEFT JOIN m ON m.lga = r.lga
        ORDER BY r.refusals DESC, r.lga
    """), params_lg)
    # Estimate missed children per LGA using the same ratio as the round
    # total. The per-LGA `missed_individuals` from the SQL is always 0
    # (refusing households never enter the individual repeat-group) — see
    # `estimated_missed_individuals` instead.
    avg_kids = next((rd["avg_children_per_household"] for rd in rounds_rows
                     if rd.get("round_number") and rd["total_forms"] > 0
                     and rd.get("round_number") == max(
                         (rr["round_number"] for rr in rounds_rows if rr.get("round_number")),
                         default=None)), 0)
    by_lga = [
        {"lga": r.lga, "total_forms": int(r.total_forms), "refusals": int(r.refusals),
         "estimated_missed_individuals": int(round(int(r.refusals) * avg_kids)),
         "refusal_pct": float(r.refusal_pct or 0)}
        for r in lga_res.fetchall()
    ]

    return {
        "rounds": rounds_rows,
        "by_reason": by_reason,
        "free_text": free_text,
        "by_lga": by_lga,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/coverage/lga-by-age
# Per-LGA coverage broken down by 1-11 months vs 12-59 months. Drives the
# "Coverage by Age Category" card on the Coverage Analysis page. Baselines
# come from the R5 target file (target_1_11_*, target_12_59_*); treated counts
# come from mda_individuals filtered to the LGA + treatment_status='1'.
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/coverage/lga-by-age")
async def coverage_lga_by_age(
    lga: Optional[str] = None,
    ward: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    params: dict = {"pid": pid}
    lgas = allowed_lgas_of(_u)
    lga_b = _lga_and(lgas, "lga", params)
    lga_h = _lga_and(lgas, "h.lga", params)
    # Optional drill-down filters so the age chart re-scopes with the page.
    f_b = ""; f_h = ""
    if lga:
        f_b += " AND UPPER(TRIM(lga)) = UPPER(TRIM(:lga))"
        f_h += " AND UPPER(TRIM(h.lga)) = UPPER(TRIM(:lga))"
        params["lga"] = lga
    if ward:
        f_b += " AND UPPER(TRIM(ward)) = UPPER(TRIM(:ward))"
        f_h += " AND UPPER(TRIM(h.ward_name)) = UPPER(TRIM(:ward))"
        params["ward"] = ward
    res = await db.execute(text(f"""
        WITH baseline AS (
          SELECT INITCAP(TRIM(lga)) AS lga,
                 SUM(COALESCE(target_1_11_f,  0) + COALESCE(target_1_11_m,  0)) AS bl_1_11,
                 SUM(COALESCE(target_12_59_f, 0) + COALESCE(target_12_59_m, 0)) AS bl_12_59
          FROM mda_baseline
          WHERE project_id = :pid{lga_b}{f_b}
          GROUP BY INITCAP(TRIM(lga))
        ),
        treated AS (
          SELECT INITCAP(TRIM(h.lga)) AS lga,
                 COUNT(*) FILTER (WHERE i.age_in_months BETWEEN 1 AND 11
                                    AND i.treatment_status = '1') AS tr_1_11,
                 COUNT(*) FILTER (WHERE i.age_in_months BETWEEN 12 AND 59
                                    AND i.treatment_status = '1') AS tr_12_59
          FROM mda_individuals i
          JOIN mda_households h ON h.formid = i.hh_formid
          WHERE h.project_id = :pid AND i.project_id = :pid
            AND i.age_in_months IS NOT NULL
            AND i.age_in_months BETWEEN 1 AND 59
            AND h.lga IS NOT NULL{lga_h}{f_h}
          GROUP BY INITCAP(TRIM(h.lga))
        )
        SELECT b.lga,
               COALESCE(b.bl_1_11,   0) AS bl_1_11,
               COALESCE(b.bl_12_59,  0) AS bl_12_59,
               COALESCE(t.tr_1_11,   0) AS tr_1_11,
               COALESCE(t.tr_12_59,  0) AS tr_12_59
        FROM baseline b
        LEFT JOIN treated t ON t.lga = b.lga
        ORDER BY b.lga
    """), params)
    rows = []
    for r in res.fetchall():
        bl1, bl2, t1, t2 = int(r.bl_1_11 or 0), int(r.bl_12_59 or 0), int(r.tr_1_11 or 0), int(r.tr_12_59 or 0)
        rows.append({
            "lga": r.lga,
            "baseline_1_11":  bl1,
            "treated_1_11":   t1,
            "coverage_1_11_pct":  round(100.0 * t1 / bl1, 1) if bl1 > 0 else 0.0,
            "baseline_12_59": bl2,
            "treated_12_59":  t2,
            "coverage_12_59_pct": round(100.0 * t2 / bl2, 1) if bl2 > 0 else 0.0,
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/qc/heatmap-geojson
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/qc/heatmap-geojson")
async def qc_heatmap_geojson(
    flag: str = "all",
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """GeoJSON of flagged GPS points for heatmap rendering in geo view."""
    where_map = {
        "outside_lga":  "flag_gps_outside_lga = TRUE",
        "outside_ward": "flag_gps_outside_ward = TRUE",
        "duplicate":    "flag_duplicate_gps = TRUE",
        "all": "(flag_gps_outside_lga = TRUE OR flag_gps_outside_ward = TRUE OR flag_duplicate_gps = TRUE)",
    }
    where = where_map.get(flag, where_map["all"])
    params: dict = {"pid": pid}
    lga_sql = _lga_and(allowed_lgas_of(_u), "lga", params)
    result = await db.execute(text(f"""
        SELECT latitude, longitude, lga, hq_user,
               CASE WHEN flag_duplicate_gps      THEN 'duplicate'
                    WHEN flag_gps_outside_lga    THEN 'outside_lga'
                    WHEN flag_gps_outside_ward   THEN 'outside_ward'
                    ELSE 'other' END AS flag_type
        FROM mda_households
        WHERE project_id = :pid
          AND latitude IS NOT NULL AND longitude IS NOT NULL
          AND latitude  BETWEEN 4 AND 14
          AND longitude BETWEEN 2 AND 15
          AND {where}{lga_sql}
        LIMIT 15000
    """), params)
    rows = result.fetchall()
    features = [
        {
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [float(r.longitude), float(r.latitude)]},
            "properties": {"lga": r.lga, "hq_user": r.hq_user, "flag_type": r.flag_type},
        }
        for r in rows
    ]
    return {"type": "FeatureCollection", "features": features}


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/qc/stacked-points  — clustered duplicate-GPS points as GeoJSON
#
# Groups every household flagged with duplicate_gps by its (lat, lon) so each
# feature returned represents a single "stack" of forms sharing an identical
# coordinate. The properties include the count and lightweight per-form
# details so the map popup can show WHY that pin was flagged (how many forms,
# from which teams, over what date range).
#
# Optional filters (hq_user / lga / ward) let the frontend jump straight from
# a team's Stacked Points cell to a map showing only that team's clusters.
# ─────────────────────────────────────────────────────────────────────────────


@router.get("/qc/stacked-points")
async def qc_stacked_points(
    hq_user: Optional[str] = None,
    lga:     Optional[str] = None,
    ward:    Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Aggregated duplicate-GPS clusters — one feature per (lat, lon) group.

    Feature properties:
      * ``stack_count``      — number of forms at that coordinate
      * ``team_list``        — comma-separated hq_user set (top 5)
      * ``teams_count``      — total distinct teams at the coord
      * ``lga``              — LGA(s) reported at that stack (typically one)
      * ``ward_names``       — ward names(s) reported
      * ``first_seen``       — earliest received_on in local time
      * ``last_seen``        — latest received_on in local time
      * ``sample_formids``   — up to 5 formids for spot-checking
    """
    # Two-stage logic when hq_user is set:
    #   1. Find every (lat, lon) where THIS TEAM has at least one flagged form.
    #   2. Return the full cluster at each of those coords — even if the OTHER
    #      forms in the cluster belong to different teams. That's the useful
    #      signal ("this team is stacking with others") rather than "this
    #      team is stacking only against itself" (which would miss cases
    #      where a rogue worker sits at one spot alongside a colleague).
    # Without hq_user, the filter is just "all duplicate-GPS forms".
    def _build_filters(alias: str) -> str:
        prefix = f"{alias}." if alias else ""
        clauses = [
            f"{prefix}project_id = :pid",
            f"{prefix}flag_duplicate_gps = TRUE",
            f"{prefix}latitude IS NOT NULL AND {prefix}longitude IS NOT NULL",
            f"{prefix}latitude  BETWEEN 4 AND 14",
            f"{prefix}longitude BETWEEN 2 AND 15",
        ]
        if lga:
            clauses.append(f"UPPER(TRIM({prefix}lga)) = UPPER(TRIM(:lga))")
        if ward:
            clauses.append(f"REPLACE(LOWER(TRIM({prefix}ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')")
        return " AND ".join(clauses)

    params: dict = {"pid": pid}
    if lga:  params["lga"]  = lga
    if ward: params["ward"] = ward
    lga_scope = _lga_and(allowed_lgas_of(_u), "lga", params)   # returns "" or " AND ..."

    team_coord_cte = ""
    join_team_coords = ""
    if hq_user:
        params["hq_user"] = hq_user
        team_scope = _lga_and(allowed_lgas_of(_u), "lga", params)  # deterministic — same lgas set
        # Note: uses no alias so the plain `_build_filters("")` clauses match.
        team_coord_cte = f"""
          WITH team_coords AS (
            SELECT DISTINCT latitude, longitude
            FROM mda_households
            WHERE {_build_filters("")} AND hq_user = :hq_user{team_scope}
          )
        """
        join_team_coords = "JOIN team_coords tc ON h.latitude = tc.latitude AND h.longitude = tc.longitude"

    result = await db.execute(text(f"""
        {team_coord_cte}
        SELECT
          h.latitude, h.longitude,
          COUNT(*)                                         AS stack_count,
          COUNT(DISTINCT h.hq_user)                        AS teams_count,
          (ARRAY_AGG(DISTINCT h.hq_user))[1:5]             AS team_list,
          (ARRAY_AGG(DISTINCT h.lga))[1:3]                 AS lga_list,
          (ARRAY_AGG(DISTINCT h.ward_name))[1:3]           AS ward_list,
          MIN((h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)::text AS first_seen,
          MAX((h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)::text AS last_seen,
          (ARRAY_AGG(h.formid))[1:5]                       AS sample_formids
        FROM mda_households h
        {join_team_coords}
        WHERE {_build_filters("h")}{lga_scope}
        GROUP BY h.latitude, h.longitude
        HAVING COUNT(*) > 1
        ORDER BY stack_count DESC
        LIMIT 5000
    """), params)

    rows = result.fetchall()
    features = [
        {
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [float(r.longitude), float(r.latitude)]},
            "properties": {
                "stack_count":    int(r.stack_count),
                "teams_count":    int(r.teams_count or 0),
                "team_list":      [t for t in (r.team_list or []) if t],
                "lga":            (r.lga_list[0] if r.lga_list else None),
                "lga_list":       [x for x in (r.lga_list or []) if x],
                "ward_names":     [x for x in (r.ward_list or []) if x],
                "first_seen":     r.first_seen,
                "last_seen":      r.last_seen,
                "sample_formids": [x for x in (r.sample_formids or []) if x],
            },
        }
        for r in rows
    ]
    return {"type": "FeatureCollection", "features": features}


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/qc/team-stacked-trend  — daily stacked-point count per team
#
# Feeds the team-trend line chart shown when the operator clicks a team's
# Stacked Points cell. Returns one row per date within the project's window
# so the chart is dense (zero-days included) — Chart.js just plots them.
# ─────────────────────────────────────────────────────────────────────────────


@router.get("/qc/team-stacked-trend")
async def qc_team_stacked_trend(
    hq_user: str,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Cumulative stacked-point count over time for a single team.

    * X axis: date (received_on in local time)
    * Y axis: cumulative number of that team's forms flagged flag_duplicate_gps

    Cumulative rather than daily so the line only ever goes up — makes it
    obvious at a glance whether a team's stacked-point problem is growing
    or plateauing.
    """
    lgas_scope = allowed_lgas_of(_u)
    params: dict = {"pid": pid, "hq_user": hq_user}
    lga_sql = _lga_and(lgas_scope, "lga", params)
    result = await db.execute(text(f"""
        WITH daily AS (
          SELECT
            (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date AS d,
            COUNT(*) FILTER (WHERE flag_duplicate_gps = TRUE) AS stacked_daily,
            COUNT(*)                                          AS forms_daily
          FROM mda_households
          WHERE project_id = :pid
            AND hq_user   = :hq_user
            AND received_on IS NOT NULL
            {lga_sql}
          GROUP BY 1
        )
        SELECT
          d::text                              AS date,
          stacked_daily,
          forms_daily,
          SUM(stacked_daily) OVER (ORDER BY d) AS stacked_cumulative,
          SUM(forms_daily)   OVER (ORDER BY d) AS forms_cumulative
        FROM daily
        ORDER BY d
    """), params)
    return [
        {
            "date":                r.date,
            "stacked_daily":       int(r.stacked_daily or 0),
            "forms_daily":         int(r.forms_daily or 0),
            "stacked_cumulative":  int(r.stacked_cumulative or 0),
            "forms_cumulative":    int(r.forms_cumulative or 0),
        }
        for r in result.fetchall()
    ]


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/teams/movement-geojson
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/teams/movement-geojson")
async def teams_movement_geojson(
    hq_user: Optional[str] = None,
    date: Optional[str] = None,
    lga: Optional[str] = None,
    ward: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Timestamped GPS points for team movement visualisation.

    Optional filters narrow the result to a single team, day, LGA or ward —
    used by the Geographic View "Team Movement" overlay when the user clicks
    a ward/settlement to inspect movement in that area.
    """
    params: dict = {"pid": pid}
    filters = [
        "project_id = :pid",
        "latitude IS NOT NULL", "longitude IS NOT NULL",
        "latitude BETWEEN 4 AND 14", "longitude BETWEEN 2 AND 15",
        "started_time IS NOT NULL",
    ]
    if hq_user:
        filters.append("hq_user = :hq_user")
        params["hq_user"] = hq_user
    if date:
        # Use CAST(...) rather than `::date` — when this WHERE clause is
        # embedded inside a CTE, SQLAlchemy's text-parameter tokenizer
        # mis-reads the consecutive colons as two named params. Bind the
        # date as a real `date` object so asyncpg accepts the DATE param.
        try:
            from datetime import date as _date_t
            params["date"] = _date_t.fromisoformat(date)
        except (TypeError, ValueError):
            raise HTTPException(400, f"Invalid date format: {date!r} (expected YYYY-MM-DD)")
        filters.append("DATE(started_time AT TIME ZONE 'UTC' + INTERVAL '1 hour') = :date")
    if lga:
        filters.append("UPPER(TRIM(lga)) = UPPER(TRIM(:lga))")
        params["lga"] = lga
    if ward:
        filters.append("UPPER(TRIM(ward_name)) = UPPER(TRIM(:ward))")
        params["ward"] = ward
    # Resolve the state's boundary project so we can spatially-tag each
    # movement point with the settlement (or ward) polygon it sits in. The
    # frontend uses settlement_uniq_cod to break the movement line whenever
    # a team crosses a settlement boundary — so each segment of the line
    # represents a single team's movement *within* one settlement.
    bres = await db.execute(text("""
        SELECT MIN(p2.id)
        FROM geo_projects p1
        JOIN geo_projects p2 ON p2.state_name = p1.state_name
        WHERE p1.id = :pid
          AND EXISTS (SELECT 1 FROM settlements s WHERE s.project_id = p2.id)
    """), {"pid": pid})
    brow = bres.fetchone()
    boundary_pid = (brow[0] if brow and brow[0] else pid)
    params["boundary_pid"] = boundary_pid
    where = " AND ".join(filters) + _lga_and(allowed_lgas_of(_u), "lga", params)
    result = await db.execute(text(f"""
        WITH base AS (
            SELECT id, latitude, longitude, hq_user, lga, ward_name, started_time,
                   ST_SetSRID(ST_MakePoint(longitude, latitude), 4326) AS geom
            FROM mda_households
            WHERE {where}
        )
        SELECT b.latitude, b.longitude, b.hq_user, b.lga, b.ward_name, b.started_time,
               EXTRACT(HOUR FROM b.started_time AT TIME ZONE 'UTC' + INTERVAL '1 hour') AS local_hour,
               (SELECT s.unique_cod FROM settlements s
                  WHERE s.project_id = :boundary_pid AND ST_Within(b.geom, s.geom)
                  LIMIT 1) AS settlement_uniq_cod,
               (SELECT s.settlement_name FROM settlements s
                  WHERE s.project_id = :boundary_pid AND ST_Within(b.geom, s.geom)
                  LIMIT 1) AS settlement_name
        FROM base b
        ORDER BY b.hq_user, b.started_time
        LIMIT 20000
    """), params)
    rows = result.fetchall()
    features = [
        {
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [float(r.longitude), float(r.latitude)]},
            "properties": {
                "hq_user":             r.hq_user,
                "lga":                 r.lga,
                "ward_name":           r.ward_name,
                "started_time":        r.started_time.isoformat() if r.started_time else None,
                "local_hour":          int(r.local_hour) if r.local_hour is not None else None,
                "settlement_uniq_cod": r.settlement_uniq_cod,
                "settlement_name":     r.settlement_name,
            },
        }
        for r in rows
    ]
    return {"type": "FeatureCollection", "features": features}


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/campaign-dates  — dataset date boundaries for filter inputs
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/campaign-dates")
async def campaign_dates(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Returns min/max received_on dates plus list of all distinct dates."""
    result = await db.execute(text("""
        SELECT
          MIN(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date AS min_date,
          MAX(received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date AS max_date,
          array_agg(DISTINCT (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
                    ORDER BY (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date) AS dates
        FROM mda_households
        WHERE project_id = :pid AND received_on IS NOT NULL
          -- Only campaign days: drop pre-campaign (test/stray) submissions before
          -- the official start date, so the date filter starts on day 1 (the day
          -- the campaign started), not an earlier stray-submission day.
          AND (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date
              >= COALESCE((SELECT campaign_start_date FROM geo_projects WHERE id = :pid), '1900-01-01'::date)
    """), {"pid": pid})
    row = result.fetchone()
    if not row or not row[0]:
        return {"min_date": None, "max_date": None, "dates": []}
    return {
        "min_date": str(row[0]),
        "max_date": str(row[1]),
        "dates":    [str(d) for d in (row[2] or [])],
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/geo/completeness
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/geo/settlement-breakdown")
async def geo_settlement_breakdown(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Settlement coverage breakdown, per-LGA + grand-total.

    Buckets a settlement into one of three states:

      * **Not visited**       — no GPS point has ever been logged inside it.
      * **Partially covered** — at least one point but grid completeness <70%.
      * **Fully covered**     — `is_visited` (≥70% grid completeness, or any
        point inside the settlement polygon when no grid is loaded).

    The 70% threshold mirrors the rest of the dashboard's "covered" rule —
    consistent with `/geo/completeness` and the Geographic View map shading.
    Used by the Settlement Coverage Breakdown table on the Geographic View
    so supervisors can see at a glance how many settlements per LGA still
    need any team contact vs which need a second pass to fill the grid.
    """
    params: dict = {"pid": pid}
    lga_sql = _lga_and(allowed_lgas_of(_u), "sa.lga_name", params)
    res = await db.execute(text(f"""
        WITH s AS (
          SELECT
            sa.lga_name,
            CASE
              WHEN COALESCE(sa.point_count, 0) = 0 THEN 'not_visited'
              WHEN sa.is_visited OR COALESCE(sa.completeness_pct, 0) >= 70 THEN 'fully_covered'
              ELSE 'partially_covered'
            END AS bucket
          FROM settlement_analytics sa
          WHERE sa.project_id = :pid{lga_sql}
        )
        SELECT
          COALESCE(lga_name, '—')                              AS lga,
          COUNT(*)                                              AS total,
          COUNT(*) FILTER (WHERE bucket = 'not_visited')        AS not_visited,
          COUNT(*) FILTER (WHERE bucket = 'partially_covered')  AS partially_covered,
          COUNT(*) FILTER (WHERE bucket = 'fully_covered')      AS fully_covered
        FROM s
        GROUP BY lga_name
        ORDER BY lga_name NULLS LAST
    """), params)
    rows = res.fetchall()
    per_lga = []
    tot_total = tot_not = tot_partial = tot_full = 0
    for r in rows:
        total       = int(r.total or 0)
        not_visited = int(r.not_visited or 0)
        partial     = int(r.partially_covered or 0)
        full        = int(r.fully_covered or 0)
        per_lga.append({
            "lga":                 r.lga,
            "total":               total,
            "not_visited":         not_visited,
            "partially_covered":   partial,
            "fully_covered":       full,
            "covered_pct":         round(100.0 * full / total, 1) if total else 0.0,
        })
        tot_total   += total
        tot_not     += not_visited
        tot_partial += partial
        tot_full    += full
    return {
        "total":               tot_total,
        "not_visited":         tot_not,
        "partially_covered":   tot_partial,
        "fully_covered":       tot_full,
        "covered_pct":         round(100.0 * tot_full / tot_total, 1) if tot_total else 0.0,
        "per_lga":             per_lga,
    }


@router.get("/teams/footprint")
async def teams_footprint(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Team-centric counterpart to /teams/by-lga.

    One row per CommCare HQ user, showing their LGA footprint and quality
    profile. Lets supervisors see at a glance which teams roam multiple
    LGAs vs which stay fixed, and which teams have the highest QC issues.

    Columns:
      team           = hq_user (or 'Unknown' when null)
      primary_lga    = LGA where the team submitted the most forms
      lgas_touched   = COUNT(DISTINCT lga)
      forms          = total household forms
      days_active    = COUNT(DISTINCT submission_date)
      avg_per_day    = forms ÷ days_active
      quality_pct    = 100 - (after_hours + fast_forms) ÷ forms × 100, clamped 0-100
      last_seen      = ISO date string
    """
    params: dict = {"pid": pid}
    lga_sql = _lga_and(allowed_lgas_of(_u), "lga", params)
    res = await db.execute(text(f"""
        WITH per_team_lga AS (
          SELECT
            COALESCE(hq_user, 'Unknown') AS team,
            INITCAP(LOWER(TRIM(lga)))    AS lga,
            COUNT(*)                     AS forms
          FROM mda_households
          WHERE project_id = :pid AND lga IS NOT NULL AND TRIM(lga) <> ''{lga_sql}
          GROUP BY 1, 2
        ),
        primary_lga AS (
          SELECT DISTINCT ON (team) team, lga AS primary_lga, forms AS primary_forms
          FROM per_team_lga
          ORDER BY team, forms DESC, lga
        ),
        per_team AS (
          SELECT
            COALESCE(hq_user, 'Unknown') AS team,
            COUNT(*) AS forms,
            COUNT(DISTINCT lga) FILTER (WHERE lga IS NOT NULL) AS lgas_touched,
            COUNT(DISTINCT (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date) AS days_active,
            SUM(CASE WHEN flag_after_hours THEN 1 ELSE 0 END) AS after_hours,
            SUM(CASE WHEN flag_fast_form   THEN 1 ELSE 0 END) AS fast_forms,
            MAX((received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)::text AS last_seen
          FROM mda_households
          WHERE project_id = :pid{lga_sql}
          GROUP BY COALESCE(hq_user, 'Unknown')
        )
        SELECT
          pt.team,
          pl.primary_lga,
          pt.lgas_touched,
          pt.forms,
          pt.days_active,
          pt.after_hours,
          pt.fast_forms,
          pt.last_seen
        FROM per_team pt
        LEFT JOIN primary_lga pl ON pl.team = pt.team
        ORDER BY pt.forms DESC, pt.team
    """), params)
    out = []
    for r in res.fetchall():
        forms = int(r.forms or 0)
        days  = int(r.days_active or 0)
        bad   = (int(r.after_hours or 0) + int(r.fast_forms or 0))
        avg   = (forms / days) if days > 0 else 0.0
        q     = 100.0 - min(100.0, (bad / max(forms, 1)) * 100.0)
        out.append({
            "team":          r.team,
            "primary_lga":   r.primary_lga,
            "lgas_touched":  int(r.lgas_touched or 0),
            "forms":         forms,
            "days_active":   days,
            "avg_per_day":   round(avg, 1),
            "quality_pct":   round(q, 1),
            "last_seen":     r.last_seen,
        })
    return out


@router.get("/geo/mop-up-shortlist")
async def geo_mop_up_shortlist(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Ranked list of LGAs that need attention, with a one-line recommendation.

    Combines three signals:
      * Settlements not visited        — count of settlements with no GPS point.
      * Treatment coverage gap         — baseline_total minus children treated.
      * Field team density             — distinct teams currently working the LGA.

    Each LGA gets an urgency score = (not_visited × 2) + max(0, gap_pct × not_visited / 5)
    so LGAs with both a big un-reached-settlement count *and* a coverage shortfall
    rise first. Returns the top 10 plus a tiny recommended action string for each.
    """
    params: dict = {"pid": pid}
    lgas = allowed_lgas_of(_u)
    lga_sa = _lga_and(lgas, "sa.lga_name", params)
    lga_l = _lga_and(lgas, "lga", params)
    res = await db.execute(text(f"""
        WITH s AS (
          SELECT
            INITCAP(LOWER(TRIM(sa.lga_name))) AS lga,
            COUNT(*) FILTER (WHERE COALESCE(sa.point_count, 0) = 0) AS not_visited,
            COUNT(*) AS total_settlements
          FROM settlement_analytics sa
          WHERE sa.project_id = :pid AND sa.lga_name IS NOT NULL{lga_sa}
          GROUP BY INITCAP(LOWER(TRIM(sa.lga_name)))
        ),
        cov AS (
          SELECT
            INITCAP(LOWER(TRIM(lga))) AS lga,
            COALESCE(SUM(total_treated), 0) AS baseline_total
          FROM mda_baseline
          WHERE project_id = :pid AND lga IS NOT NULL{lga_l}
          GROUP BY INITCAP(LOWER(TRIM(lga)))
        ),
        hh AS (
          SELECT
            INITCAP(LOWER(TRIM(lga))) AS lga,
            COUNT(DISTINCT hq_user) AS teams,
            COALESCE(SUM(number_of_treated), 0) AS treated
          FROM mda_households
          WHERE project_id = :pid AND lga IS NOT NULL{lga_l}
          GROUP BY INITCAP(LOWER(TRIM(lga)))
        )
        SELECT
          COALESCE(s.lga, hh.lga, cov.lga) AS lga,
          COALESCE(s.not_visited, 0)       AS not_visited,
          COALESCE(s.total_settlements, 0) AS total_settlements,
          COALESCE(cov.baseline_total, 0)  AS baseline_total,
          COALESCE(hh.treated, 0)          AS treated,
          COALESCE(hh.teams, 0)            AS teams
        FROM s
        FULL OUTER JOIN cov ON cov.lga = s.lga
        FULL OUTER JOIN hh  ON hh.lga  = COALESCE(s.lga, cov.lga)
    """), params)
    rows = []
    for r in res.fetchall():
        if not r.lga:
            continue
        baseline = int(r.baseline_total or 0)
        treated  = int(r.treated or 0)
        nv       = int(r.not_visited or 0)
        tot_st   = int(r.total_settlements or 0)
        teams    = int(r.teams or 0)
        gap_pct  = max(0.0, 100.0 - (100.0 * treated / baseline)) if baseline else 0.0
        urgency  = (nv * 2) + max(0.0, gap_pct * nv / 5.0)
        if nv == 0 and gap_pct < 5:
            continue  # nothing to recommend here
        # One-line recommended action — pick the leading issue.
        if nv > 0 and (tot_st == 0 or nv / max(tot_st, 1) >= 0.5):
            action = f"Deploy a first-pass team — {nv} of {tot_st} settlements not yet visited."
        elif nv > 0:
            action = f"Mop-up: {nv} settlements still un-visited; current team count {teams}."
        elif gap_pct >= 20:
            action = f"Coverage gap {gap_pct:.0f}% — review baseline accuracy or expand teams."
        else:
            action = f"Top up: small coverage gap of {gap_pct:.0f}% remaining."
        rows.append({
            "lga":               r.lga,
            "not_visited":       nv,
            "total_settlements": tot_st,
            "baseline_total":    baseline,
            "treated":           treated,
            "gap_pct":           round(gap_pct, 1),
            "teams":             teams,
            "urgency":           round(urgency, 1),
            "action":            action,
        })
    rows.sort(key=lambda x: x["urgency"], reverse=True)
    return rows[:10]


@router.get("/teams/by-lga")
async def teams_by_lga(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """How many field teams are operating in each LGA, and how active.

    A "team" here is a unique CommCare HQ user (mda_households.hq_user) —
    same identity used by Teams Performance, /overview's teams_active, etc.
    Returned columns:
      teams          = distinct HQ users that submitted from this LGA
      forms          = total household forms from this LGA
      days_active    = distinct submission dates with at least one form
      avg_per_team_per_day = forms ÷ (teams × days_active), clamped
      first_seen / last_seen = ISO date strings, Africa/Lagos-anchored

    Used by the Teams per LGA decision-support table on Coverage Analysis
    when supervisors decide where to redeploy teams for mop-up. Scoped by
    project_id so each round/state stays in its own scope.
    """
    # Pull the LGA-level activity in one query, then join settlement counts
    # so we can derive a target-team benchmark per LGA. The benchmark uses
    # the SARMAAN rate (80 forms / team / day) — if an LGA has N settlements,
    # and one settlement ≈ one form, then expected_teams ≈ N ÷ (80 × planned_days)
    # so a team can finish in the campaign window. Falls back gracefully when
    # planned days isn't set.
    params: dict = {"pid": pid}
    lgas = allowed_lgas_of(_u)
    lga_l = _lga_and(lgas, "lga", params)
    lga_sa = _lga_and(lgas, "lga_name", params)
    res = await db.execute(text(f"""
        WITH hh AS (
          SELECT
            INITCAP(LOWER(TRIM(lga)))                       AS lga,
            COUNT(DISTINCT hq_user)                         AS teams,
            COUNT(*)                                        AS forms,
            COUNT(DISTINCT (received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date) AS days_active,
            MIN((received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)::text     AS first_seen,
            MAX((received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)::text     AS last_seen
          FROM mda_households
          WHERE project_id = :pid AND lga IS NOT NULL AND TRIM(lga) <> ''{lga_l}
          GROUP BY INITCAP(LOWER(TRIM(lga)))
        ),
        st AS (
          SELECT INITCAP(LOWER(TRIM(lga_name))) AS lga,
                 COUNT(*) AS settlements
          FROM settlement_analytics
          WHERE project_id = :pid AND lga_name IS NOT NULL{lga_sa}
          GROUP BY INITCAP(LOWER(TRIM(lga_name)))
        )
        SELECT hh.*, COALESCE(st.settlements, 0) AS settlements
        FROM hh LEFT JOIN st ON st.lga = hh.lga
        ORDER BY hh.teams DESC, hh.forms DESC
    """), params)

    # Planned campaign window — drives the target-team denominator.
    proj_res = await db.execute(text("""
        SELECT campaign_start_date, campaign_end_date
        FROM geo_projects WHERE id = :pid
    """), {"pid": pid})
    proj_row = proj_res.fetchone()
    planned_days = None
    if proj_row and proj_row[0] and proj_row[1]:
        planned_days = (proj_row[1] - proj_row[0]).days + 1

    out = []
    for r in res.fetchall():
        teams = int(r.teams or 0)
        forms = int(r.forms or 0)
        days  = int(r.days_active or 0)
        settl = int(r.settlements or 0)
        avg   = (forms / (teams * days)) if teams > 0 and days > 0 else 0.0
        # Target teams = settlements ÷ (80 forms/team/day × planned_days).
        # Round up so partial teams count. Falls back to None when planned
        # days isn't configured — the frontend will hide the benchmark column.
        target_teams = None
        deployment_status = None
        if planned_days and settl > 0:
            target_teams = max(1, -(-settl // (80 * planned_days)))  # ceil
            if teams == 0:
                deployment_status = "unstaffed"
            elif teams < 0.5 * target_teams:
                deployment_status = "under-staffed"
            elif teams > 1.5 * target_teams:
                deployment_status = "over-staffed"
            else:
                deployment_status = "on-target"
        out.append({
            "lga":                   r.lga,
            "teams":                 teams,
            "target_teams":          target_teams,
            "deployment_status":     deployment_status,
            "settlements":           settl,
            "forms":                 forms,
            "days_active":           days,
            "avg_per_team_per_day":  round(avg, 1),
            "first_seen":            r.first_seen,
            "last_seen":             r.last_seen,
        })
    return out


@router.get("/geo/coverage-summary")
async def geo_coverage_summary(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """High-level coverage roll-up across all three admin levels.

    Returns three rows — LGA / Ward / Settlement — each showing how many
    units fall at or above the 70% coverage threshold vs below it. An
    LGA / Ward is "at threshold" when ≥70% of its settlements are fully
    covered. Settlements use their own is_visited / completeness rule
    (≥70% grid coverage, or any GPS point inside when grids are absent).

    This is the operator-facing decision-support table: at a glance, how
    many LGAs and Wards are still in the "need attention" bucket, and how
    many settlements have not been reached.
    """
    # First compute the per-settlement bucket once and reuse for LGA + Ward
    # rollups. Saves us doing three near-identical CTEs.
    params: dict = {"pid": pid}
    lga_sql = _lga_and(allowed_lgas_of(_u), "sa.lga_name", params)
    s_res = await db.execute(text(f"""
        WITH s AS (
          SELECT sa.lgacode, sa.lga_name, sa.wardcode, sa.ward_name,
                 CASE WHEN sa.is_visited OR COALESCE(sa.completeness_pct, 0) >= 70
                      THEN 1 ELSE 0 END AS covered
          FROM settlement_analytics sa
          WHERE sa.project_id = :pid{lga_sql}
        ),
        per_lga AS (
          SELECT lgacode, AVG(covered) AS frac
          FROM s
          GROUP BY lgacode
        ),
        per_ward AS (
          SELECT wardcode, AVG(covered) AS frac
          FROM s
          GROUP BY wardcode
        )
        SELECT
          (SELECT COUNT(*)                       FROM per_lga)  AS lga_total,
          (SELECT COUNT(*) FILTER (WHERE frac >= 0.7) FROM per_lga)  AS lga_at,
          (SELECT COUNT(*)                       FROM per_ward) AS ward_total,
          (SELECT COUNT(*) FILTER (WHERE frac >= 0.7) FROM per_ward) AS ward_at,
          (SELECT COUNT(*)                       FROM s)        AS sett_total,
          (SELECT COUNT(*) FILTER (WHERE covered = 1) FROM s)   AS sett_at
    """), params)
    row = s_res.fetchone()
    def split(total, atv):
        total = int(total or 0); atv = int(atv or 0)
        below = total - atv
        return {
            "total":      total,
            "at_target":  atv,
            "below":      below,
            "pct":        round(100.0 * atv / total, 1) if total else 0.0,
        }
    if not row or not int(row.sett_total or 0):
        return {
            "threshold": 70,
            "lga":         split(0, 0),
            "ward":        split(0, 0),
            "settlement":  split(0, 0),
        }
    return {
        "threshold":   70,
        "lga":         split(row.lga_total,  row.lga_at),
        "ward":        split(row.ward_total, row.ward_at),
        "settlement":  split(row.sett_total, row.sett_at),
    }


@router.get("/geo/completeness")
async def geo_completeness(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Overall grid completeness for the active round.

    overall_completeness = mean of per-settlement completeness with the
    ≥70 % round-up rule applied (so a settlement at 75 % counts as 100 %
    in the average). visited_settlements is the count of settlements with
    ≥1 GPS point inside their polygon (no completeness gate).
    """
    params: dict = {"pid": pid}
    lga_sql = _lga_and(allowed_lgas_of(_u), "sa.lga_name", params)
    result = await db.execute(text(f"""
        SELECT
          ROUND(AVG(CASE WHEN COALESCE(completeness_pct, 0) >= 70
                          THEN 100.0
                          ELSE COALESCE(completeness_pct, 0) END)::numeric, 1) AS overall_completeness,
          COUNT(*) FILTER (WHERE COALESCE(completeness_pct, 0) >= 60)  AS completed_60,
          COUNT(*) FILTER (WHERE COALESCE(completeness_pct, 0) >= 70)  AS completed_70,
          COUNT(*) FILTER (WHERE COALESCE(point_count, 0)       > 0)   AS visited_settlements,
          COUNT(*) AS total_settlements
        FROM settlement_analytics sa
        WHERE sa.project_id = :pid{lga_sql}
    """), params)
    row = result.fetchone()
    if not row:
        return {"overall_completeness": 0.0, "completed_60": 0, "completed_70": 0,
                "visited_settlements": 0, "total_settlements": 0, "visitation_pct": 0.0}
    total = int(row[4] or 0)
    visited = int(row[3] or 0)
    return {
        "overall_completeness": float(row[0] or 0),
        "completed_60":         int(row[1] or 0),
        "completed_70":         int(row[2] or 0),
        "visited_settlements":  visited,
        "total_settlements":    total,
        "visitation_pct":       round(100.0 * visited / total, 1) if total else 0.0,
    }


@router.get("/geo/wards-coverage")
async def geo_wards_coverage(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Ward polygons (GeoJSON) tagged with coverage % — public aggregate, no
    PII. Drives the app's map page (loaded in a WebView) and any web map.
    Geometry comes from the state's boundary project; coverage from
    settlement_analytics for the selected project, joined by round-stable
    wardcode."""
    # Own-boundaries-first (matches the web analytics + the sync recompute) so the
    # wardcode join lines up with settlement_analytics for this round.
    from app.services.spatial_engine import _resolve_boundary_pid
    bpid = await _resolve_boundary_pid(pid, db)
    params: dict = {"pid": pid, "bpid": bpid}
    lga_sql = _lga_and(allowed_lgas_of(_u), "w.lga_name", params)
    res = await db.execute(text(f"""
        WITH cov AS (
          -- Visitation: share of a ward's settlements physically reached
          -- (a GPS visit or >=70% form completeness). This is what the map
          -- COLOURS by: "where have we actually been?" - the natural question
          -- for a map, and distinct from administrative coverage below.
          SELECT wardcode,
                 AVG(CASE WHEN is_visited OR COALESCE(completeness_pct, 0) >= 70
                          THEN 1.0 ELSE 0.0 END) AS frac
          FROM settlement_analytics
          WHERE project_id = :pid
          GROUP BY wardcode
        ),
        tgt AS (
          -- Ward-level TARGET from the baseline workbook (same source as the
          -- Coverage Analysis page's ward drill-down), keyed by ward name.
          SELECT UPPER(TRIM(mb.lga))  AS lga_key,
                 UPPER(TRIM(mb.ward)) AS ward_key,
                 SUM(mb.total_treated) AS baseline_total
          FROM mda_baseline mb
          WHERE mb.project_id = :pid AND mb.ward IS NOT NULL AND TRIM(mb.ward) <> ''
          GROUP BY 1, 2
        ),
        act AS (
          -- Actual treated per ward from submitted households.
          SELECT UPPER(TRIM(h.lga))       AS lga_key,
                 UPPER(TRIM(h.ward_name))  AS ward_key,
                 COALESCE(SUM(h.number_of_treated), 0) AS actual_treated
          FROM mda_households h
          WHERE h.project_id = :pid AND h.ward_name IS NOT NULL
          GROUP BY 1, 2
        )
        SELECT w.ward_name, w.lga_name, w.wardcode,
               ST_AsGeoJSON(w.geom) AS geom,
               COALESCE(cov.frac, 0)::float AS frac,
               tgt.baseline_total AS target,
               COALESCE(act.actual_treated, 0) AS treated
        FROM wards w
        LEFT JOIN cov ON cov.wardcode = w.wardcode
        LEFT JOIN tgt ON tgt.lga_key = UPPER(TRIM(w.lga_name))
                     AND tgt.ward_key = UPPER(TRIM(w.ward_name))
        LEFT JOIN act ON act.lga_key = UPPER(TRIM(w.lga_name))
                     AND act.ward_key = UPPER(TRIM(w.ward_name))
        WHERE w.project_id = :bpid{lga_sql}
    """), params)
    import json as _json
    features = []
    for r in res.fetchall():
        if not r.geom:
            continue
        frac = float(r.frac or 0)
        target = float(r.target) if r.target is not None else None
        treated = float(r.treated or 0)
        # Administrative coverage = treated / ward-target (matches the Coverage
        # Analysis page). None when this ward has no target in the baseline so
        # the popup can say "no target" rather than a misleading 0%.
        admin_cov = round(100.0 * treated / target, 1) if target and target > 0 else None
        features.append({
            "type": "Feature",
            "geometry": _json.loads(r.geom),
            "properties": {
                "ward_name": r.ward_name,
                "lga_name": r.lga_name,
                # coverage_pct kept as visitation for backward-compatible map
                # colouring; visitation_pct is the same value, named plainly.
                "coverage_pct": round(100.0 * frac, 1),
                "visitation_pct": round(100.0 * frac, 1),
                "admin_coverage_pct": admin_cov,
                "treated": int(treated),
                "target": int(target) if target else None,
                "is_at_target": frac >= 0.7,
            },
        })
    return {"type": "FeatureCollection", "project_id": pid, "features": features}


@router.get("/geo/lgas-coverage")
async def geo_lgas_coverage(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """LGA polygons (GeoJSON) coloured by coverage — public aggregate."""
    # Own-boundaries-first (matches the web analytics + the sync recompute).
    from app.services.spatial_engine import _resolve_boundary_pid
    bpid = await _resolve_boundary_pid(pid, db)
    params: dict = {"pid": pid, "bpid": bpid}
    lga_sql = _lga_and(allowed_lgas_of(_u), "l.lga_name", params)
    res = await db.execute(text(f"""
        WITH cov AS (
          SELECT lgacode,
                 AVG(CASE WHEN is_visited OR COALESCE(completeness_pct,0) >= 70 THEN 1.0 ELSE 0.0 END) AS frac,
                 -- visitation = share of settlements with >=1 GPS point (matches
                 -- the web map's LGA choropleth; "have we reached here at all?").
                 AVG(CASE WHEN COALESCE(point_count,0) > 0 THEN 1.0 ELSE 0.0 END) AS visit_frac
          FROM settlement_analytics WHERE project_id = :pid GROUP BY lgacode
        ),
        campaign AS (
          -- LGAs actually part of THIS round: a baseline target OR submitted
          -- households. Project-dynamic, so it's correct for any state/round
          -- (not just whichever LGAs happen to have settlement_analytics rows).
          SELECT DISTINCT upper(trim(lga)) AS lga_u FROM mda_baseline   WHERE project_id = :pid AND lga IS NOT NULL AND trim(lga) <> ''
          UNION
          SELECT DISTINCT upper(trim(lga)) AS lga_u FROM mda_households  WHERE project_id = :pid AND lga IS NOT NULL AND trim(lga) <> ''
        )
        SELECT l.lga_name, l.lgacode,
               ST_AsGeoJSON(l.geom) AS geom,
               COALESCE(cov.frac, 0)::float AS frac,
               COALESCE(cov.visit_frac, 0)::float AS visit_frac,
               EXISTS (SELECT 1 FROM campaign c WHERE c.lga_u = upper(trim(l.lga_name))) AS in_campaign
        FROM lgas l LEFT JOIN cov ON cov.lgacode = l.lgacode
        WHERE l.project_id = :bpid{lga_sql}
    """), params)
    import json as _json
    feats = []
    for r in res.fetchall():
        if not r.geom:
            continue
        frac = float(r.frac or 0)
        # in_campaign: the LGA was targeted (baseline) or submitted forms this
        # round. LGAs in the boundary but NOT in the campaign render blank on the
        # map instead of being graded red for 0% — works for any state/round.
        feats.append({"type": "Feature", "geometry": _json.loads(r.geom),
                      "properties": {"lga_name": r.lga_name, "lgacode": r.lgacode,
                                     "coverage_pct": round(100.0 * frac, 1),
                                     "visitation_pct": round(100.0 * float(r.visit_frac or 0), 1),
                                     "is_at_target": frac >= 0.7,
                                     "in_campaign": bool(r.in_campaign)}})
    return {"type": "FeatureCollection", "project_id": pid, "features": feats}


@router.get("/geo/settlements-coverage")
async def geo_settlements_coverage(
    lgacode: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Settlement polygons (GeoJSON) coloured visited/not + completeness — the
    core coverage mosaic. Public aggregate. Optional ?lgacode= to scope to one
    LGA (lighter payload)."""
    params: dict = {"pid": pid}
    where = "sa.project_id = :pid"
    if lgacode:
        where += " AND sa.lgacode = :lgacode"
        params["lgacode"] = lgacode
    where += _lga_and(allowed_lgas_of(_u), "sa.lga_name", params)
    res = await db.execute(text(f"""
        SELECT sa.settlement_name, sa.lga_name, sa.ward_name,
               COALESCE(sa.is_visited, FALSE) AS is_visited,
               COALESCE(sa.completeness_pct, 0)::float AS completeness_pct,
               -- Simplify geometry (~33 m) — settlement polygons are tiny at map
               -- scale, so this sharply cuts the GeoJSON payload (faster load on
               -- the phone) with no visible change.
               ST_AsGeoJSON(ST_SimplifyPreserveTopology(s.geom, 0.0003)) AS geom
        FROM settlement_analytics sa
        JOIN settlements s ON s.id = sa.settlement_id
        WHERE {where}
    """), params)
    import json as _json
    feats = []
    for r in res.fetchall():
        if not r.geom:
            continue
        feats.append({"type": "Feature", "geometry": _json.loads(r.geom),
                      "properties": {"settlement_name": r.settlement_name, "lga_name": r.lga_name,
                                     "ward_name": r.ward_name, "is_visited": bool(r.is_visited),
                                     "completeness_pct": round(float(r.completeness_pct or 0), 1)}})
    return {"type": "FeatureCollection", "project_id": pid, "features": feats}


@router.get("/coverage/settlement")
async def mda_coverage_settlement(
    lga: Optional[str] = None,
    ward: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """Settlements within a ward — name, visited flag, completeness %, points.
    Lightweight (no geometry); drives the dashboard charts' ward → settlement
    drill. Filtered by LGA + ward NAME, and LGA-scoped for restricted accounts."""
    clauses = ["sa.project_id = :pid"]
    params: dict = {"pid": pid}
    # Match space/case-insensitively so the canonical ward/LGA names the charts
    # pass down (e.g. "Yadakunya") still hit settlement_analytics regardless of
    # how the boundary data cased/spaced them ("YADA KUNYA").
    if lga:
        clauses.append("REPLACE(LOWER(TRIM(sa.lga_name)), ' ', '') = REPLACE(LOWER(TRIM(:lga)), ' ', '')")
        params["lga"] = lga
    if ward:
        clauses.append("REPLACE(LOWER(TRIM(sa.ward_name)), ' ', '') = REPLACE(LOWER(TRIM(:ward)), ' ', '')")
        params["ward"] = ward
    where = " AND ".join(clauses) + _lga_and(allowed_lgas_of(_u), "sa.lga_name", params)
    res = await db.execute(text(f"""
        SELECT sa.settlement_name, sa.ward_name, sa.lga_name,
               COALESCE(sa.is_visited, FALSE) AS is_visited,
               COALESCE(sa.completeness_pct, 0)::float AS completeness_pct,
               COALESCE(sa.point_count, 0) AS point_count
        FROM settlement_analytics sa
        WHERE {where}
        ORDER BY sa.completeness_pct DESC, sa.settlement_name
    """), params)
    return [
        {"settlement_name": r.settlement_name, "ward_name": r.ward_name, "lga_name": r.lga_name,
         "is_visited": bool(r.is_visited), "completeness_pct": round(float(r.completeness_pct or 0), 1),
         "point_count": int(r.point_count or 0)}
        for r in res.fetchall()
    ]


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/settlement-status/download
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/settlement-status/download")
async def download_settlement_status(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Download Excel file for the selected round: LGA, Ward, Settlement,
    Visited (Yes/No), Completeness %. Admin/superadmin only."""
    result = await db.execute(text("""
        SELECT
            sa.lga_name,
            sa.ward_name,
            sa.settlement_name,
            CASE WHEN sa.is_visited THEN 'Visited' ELSE 'Not Visited' END AS visit_status,
            ROUND(sa.completeness_pct::numeric, 1) AS completeness_pct,
            sa.visited_grids,
            sa.total_grids,
            sa.point_count
        FROM settlement_analytics sa
        WHERE sa.project_id = :pid
        ORDER BY sa.lga_name, sa.ward_name, sa.settlement_name
    """), {"pid": pid})
    rows = result.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settlement Status"
    headers = ["LGA", "Ward", "Settlement", "Status", "Completeness %", "Grids Visited", "Total Grids", "GPS Points"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="003D1A")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    green_fill = PatternFill("solid", fgColor="DCFCE7")
    red_fill   = PatternFill("solid", fgColor="FEE2E2")
    for row_data in rows:
        ws.append(list(row_data))
        last_row = ws.max_row
        status_cell = ws.cell(row=last_row, column=4)
        status_cell.fill = green_fill if status_cell.value == "Visited" else red_fill

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"settlement_status_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Geospatial View — bulk-export downloads (admin / superadmin only).
#
# Three Excel exports mirroring the LGA → Ward → Settlement hierarchy in the
# geospatial nav panel, sharing the styling and column-width logic from the
# settlement-status download above so the team sees a consistent file shape.
# ─────────────────────────────────────────────────────────────────────────────

def _style_workbook(ws) -> None:
    """Apply the standard green header + auto-width to a worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="003D1A")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def _xlsx_response(wb, basename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{basename}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/geo/download/lgas")
async def download_geo_lgas(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Per-LGA coverage Excel for the active round (admin/superadmin only)."""
    result = await db.execute(text("""
        WITH baseline AS (
          SELECT INITCAP(LOWER(TRIM(lga))) AS lga,
                 UPPER(TRIM(lga))          AS lga_key,
                 SUM(total_treated)        AS baseline_total
          FROM mda_baseline WHERE project_id = :pid
          GROUP BY INITCAP(LOWER(TRIM(lga))), UPPER(TRIM(lga))
        )
        SELECT
          b.lga,
          b.baseline_total,
          COUNT(h.id)                                                       AS forms,
          COALESCE(SUM(h.number_of_treated), 0)                             AS treated,
          CASE WHEN b.baseline_total > 0
               THEN ROUND(100.0 * COALESCE(SUM(h.number_of_treated), 0) / b.baseline_total, 1)
               ELSE 0 END                                                    AS coverage_pct,
          COUNT(DISTINCT h.hq_user)                                         AS teams,
          COUNT(DISTINCT (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)
                                                                            AS days_active
        FROM baseline b
        LEFT JOIN mda_households h
          ON h.project_id = :pid AND UPPER(TRIM(h.lga)) = b.lga_key
        GROUP BY b.lga, b.baseline_total
        ORDER BY coverage_pct DESC, b.lga
    """), {"pid": pid})
    rows = result.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LGAs"
    ws.append(["LGA", "Baseline Target", "Forms", "Treated", "Coverage %", "Teams", "Days Active"])
    for r in rows:
        ws.append(list(r))
    _style_workbook(ws)
    return _xlsx_response(wb, "lga_coverage")


@router.get("/geo/download/wards")
async def download_geo_wards(
    lga: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Per-Ward coverage Excel for the active round (admin/superadmin only).
    Optionally narrow to a single LGA via ``?lga=…``."""
    filters: list = []
    params: dict = {"pid": pid}
    if lga:
        filters.append("UPPER(TRIM(h.lga)) = UPPER(TRIM(:lga))")
        params["lga"] = lga
    where = _scoped_where(pid, filters, params, alias="h")
    result = await db.execute(text(f"""
        SELECT
          h.lga,
          h.ward_name,
          COALESCE(b.baseline_total, 0) AS baseline_total,
          COUNT(*)                                       AS forms,
          COALESCE(SUM(h.number_of_treated), 0)          AS treated,
          CASE WHEN COALESCE(b.baseline_total, 0) > 0
               THEN ROUND(100.0 * COALESCE(SUM(h.number_of_treated), 0) / b.baseline_total, 1)
               ELSE 0 END                                AS coverage_pct,
          COUNT(DISTINCT h.hq_user)                       AS teams,
          COUNT(DISTINCT (h.received_on AT TIME ZONE 'UTC' AT TIME ZONE 'Africa/Lagos')::date)
                                                          AS days_active
        FROM mda_households h
        LEFT JOIN (
          SELECT UPPER(TRIM(lga)) AS lga, UPPER(TRIM(ward)) AS ward,
                 SUM(total_treated) AS baseline_total
          FROM mda_baseline WHERE project_id = :pid
          GROUP BY UPPER(TRIM(lga)), UPPER(TRIM(ward))
        ) b ON UPPER(TRIM(h.lga)) = b.lga AND UPPER(TRIM(h.ward_name)) = b.ward
        {where}
        GROUP BY h.lga, h.ward_name, b.baseline_total
        ORDER BY h.lga, coverage_pct DESC NULLS LAST
    """), params)
    rows = result.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wards"
    ws.append(["LGA", "Ward", "Baseline Target", "Forms", "Treated", "Coverage %", "Teams", "Days Active"])
    for r in rows:
        ws.append(list(r))
    _style_workbook(ws)
    return _xlsx_response(wb, "ward_coverage")


@router.get("/geo/download/settlements")
async def download_geo_settlements(
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _admin: User = Depends(require_admin),
):
    """Per-settlement detail Excel for the active round (admin/superadmin only).
    Adds GPS points + grid completeness alongside the visit status."""
    result = await db.execute(text("""
        SELECT
            sa.lga_name,
            sa.ward_name,
            sa.settlement_name,
            CASE WHEN sa.is_visited THEN 'Visited' ELSE 'Not Visited' END AS visit_status,
            ROUND(sa.completeness_pct::numeric, 1) AS completeness_pct,
            sa.visited_grids,
            sa.total_grids,
            sa.point_count
        FROM settlement_analytics sa
        WHERE sa.project_id = :pid
        ORDER BY sa.lga_name, sa.ward_name, sa.settlement_name
    """), {"pid": pid})
    rows = result.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settlements"
    ws.append(["LGA", "Ward", "Settlement", "Status", "Completeness %", "Grids Visited", "Total Grids", "GPS Points"])
    for r in rows:
        ws.append(list(r))
    _style_workbook(ws)
    return _xlsx_response(wb, "settlement_detail")


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/mda/wards
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/wards")
async def get_ward_list(
    lga: Optional[str] = None,
    pid: int = Depends(resolve_pid),
    db: AsyncSession = Depends(get_db),
    _u: Optional[User] = Depends(get_current_user_optional),
):
    """
    Return the real wards per LGA, anchored on the ward-level TARGETS in
    mda_baseline (the target workbook) rather than the free-text ``ward_name``
    typed on households. Household text is dirty - settlement / polling-unit
    names leak in and repeat across LGAs - so it used to list ~40 "wards" for an
    LGA that really has ~11. LGA/ward names are returned Title Case.
    - With ?lga=X  → list of ward names for that LGA
    - Without lga  → dict { lga_name: [ward1, ward2, ...] } for all LGAs
    """
    lgas = allowed_lgas_of(_u)
    if lga:
        params: dict = {"pid": pid, "lga": lga}
        lga_sql = _lga_and(lgas, "lga", params)
        result = await db.execute(text(f"""
            SELECT DISTINCT INITCAP(LOWER(TRIM(ward))) AS ward_name
            FROM mda_baseline
            WHERE project_id = :pid
              AND ward IS NOT NULL AND TRIM(ward) <> ''
              AND UPPER(TRIM(lga)) = UPPER(TRIM(:lga)){lga_sql}
            ORDER BY ward_name
        """), params)
        return [r[0] for r in result.fetchall()]
    else:
        params = {"pid": pid}
        lga_sql = _lga_and(lgas, "lga", params)
        result = await db.execute(text(f"""
            SELECT DISTINCT INITCAP(LOWER(TRIM(lga)))  AS lga,
                            INITCAP(LOWER(TRIM(ward))) AS ward_name
            FROM mda_baseline
            WHERE project_id = :pid AND ward IS NOT NULL AND TRIM(ward) <> ''
              AND lga IS NOT NULL AND TRIM(lga) <> ''{lga_sql}
            ORDER BY lga, ward_name
        """), params)
        grouped: Dict[str, list] = {}
        for row in result.fetchall():
            lga_key, ward = row[0], row[1]
            if lga_key not in grouped:
                grouped[lga_key] = []
            grouped[lga_key].append(ward)
        return grouped


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/mda/upload-mlos
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/upload-mlos")
async def upload_mlos(
    file: UploadFile = File(...),
    pid: int = Depends(resolve_pid),
    _super: User = Depends(require_superadmin),
):
    """
    Upload SARMAAN MLOS Excel file (columns: state_name, lga_name, ward_name,
    settlement_name, latitude, longitude, source).
    Each record's lat/lon is spatially joined to the wards boundary table to
    confirm/enrich ward_name from the authoritative polygon geometry.
    Replaces mlos_settlements rows for the active project only.
    """
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".csv")):
        raise HTTPException(400, "File must be .xlsx or .csv")
    raw = await file.read()

    # Read into a uniform (header_row, body_rows) shape regardless of format.
    if fname.endswith(".csv"):
        import csv as _csv
        try:
            text_blob = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text_blob = raw.decode("latin-1")
        reader   = _csv.reader(io.StringIO(text_blob))
        all_rows = list(reader)
        header   = [str(c or "").strip() for c in (all_rows[0] if all_rows else [])]
        rows_raw = [tuple(r) for r in all_rows[1:]] if len(all_rows) > 1 else []
    else:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(400, f"Cannot open workbook: {e}")
        ws       = wb[wb.sheetnames[0]]
        all_rows = list(ws.iter_rows(values_only=True))
        header   = [str(c or "").strip() for c in (all_rows[0] if all_rows else [])]
        rows_raw = all_rows[1:] if len(all_rows) > 1 else []
        wb.close()

    # Header-driven column resolution. Different partners send MLOS with
    # slightly different column names — be tolerant of underscores, case,
    # and short/long forms. Each entry below: (target_field, candidate_names).
    col_idx = _map_columns(header, {
        "state":      ("state_name", "state", "statename"),
        "lga":        ("lga_name", "lga", "lganame", "adm2_en"),
        "ward":       ("ward_name", "ward", "wardname", "adm3_en"),
        "settlement": ("settlement_name", "settlement", "settlemen", "name"),
        "lat":        ("latitude", "lat", "y_coord", "y"),
        "lon":        ("longitude", "lon", "long", "lng", "x_coord", "x"),
        "source":     ("source",),
    })

    def cell(row, field):
        idx = col_idx.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    records = []
    for row in rows_raw:
        if not row or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        state   = _str(cell(row, "state"))
        lga     = _str(cell(row, "lga"))
        ward_r  = _str(cell(row, "ward"))
        sett    = _str(cell(row, "settlement"))
        lat     = _float_safe(cell(row, "lat"))
        lon     = _float_safe(cell(row, "lon"))
        source  = _str(cell(row, "source"))
        if not lga:
            continue
        geom_wkt = f"SRID=4326;POINT({lon} {lat})" if lat and lon and not (lat == 0 and lon == 0) else None
        records.append((state, lga, ward_r, sett, lat, lon, source, geom_wkt))

    conn = _get_sync_conn()
    try:
        cur = conn.cursor()
        # Find the state's canonical boundary project (lowest-id project for the same state)
        cur.execute("""
            SELECT MIN(p2.id) FROM geo_projects p1
            JOIN geo_projects p2 ON p2.state_name = p1.state_name
            WHERE p1.id = %s
        """, (pid,))
        boundary_pid = (cur.fetchone() or [pid])[0] or pid

        cur.execute("DELETE FROM mlos_settlements WHERE project_id = %s", (pid,))
        if records:
            now_str = datetime.utcnow().isoformat()
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO mlos_settlements
                   (project_id, state_name, lga_name, ward_name_raw, settlement_name,
                    latitude, longitude, source, geom, uploaded_at)
                   VALUES %s""",
                [
                    (pid, s, l, wr, se, la, lo, src, geom, now_str)
                    for s, l, wr, se, la, lo, src, geom in records
                ],
                template="(%s,%s,%s,%s,%s,%s,%s,%s,ST_GeomFromEWKT(%s),%s)",
                page_size=500,
            )
            # Spatial join: set ward_name from the intersecting ward boundary polygon
            cur.execute("""
                UPDATE mlos_settlements m
                SET ward_name = w.ward_name
                FROM wards w
                WHERE m.project_id = %s
                  AND w.project_id = %s
                  AND m.geom IS NOT NULL
                  AND ST_Within(m.geom, w.geom)
            """, (pid, boundary_pid))
            # For records outside any boundary, fall back to ward_name_raw
            cur.execute("""
                UPDATE mlos_settlements
                SET ward_name = ward_name_raw
                WHERE project_id = %s
                  AND ward_name IS NULL AND ward_name_raw IS NOT NULL
            """, (pid,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"DB error: {e}")
    finally:
        conn.close()

    return {"rows_inserted": len(records)}


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/mda/upload-baseline
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/upload-baseline")
async def upload_baseline(
    file: UploadFile = File(...),
    pid: int = Depends(resolve_pid),
    _super: User = Depends(require_superadmin),
):
    """Upload a baseline xlsx — replaces baseline rows for the active project only.

    Auto-detects three formats (in priority order):

    1. **R5 ward target (authoritative)** — sheet named ``R5 Target (Ward)``
       (typical SARMAAN II workbook). Columns: ``S/N, LGA, Ward,
       R5 Target 1-11 Months, R5 Target 12-59 Months, R5 Target``. These are
       the growth-adjusted totals used as the coverage denominator. If a
       ``Raw Data`` sheet is also present, its female/male counts are used
       to split each age-band target proportionally between the *_f / *_m
       columns so sums still line up. If not, the target is split 50/50.

    2. **Raw Data only** — sheet named ``Raw Data`` with columns:
       ``lga, wardname, T_1_11_Months_Female, T_1_11_Months_Male,
       T_12_59_Months_Female, T_12_59_Months_Male``. ``total_treated`` is the
       direct sum of the four age/sex columns.

    3. **Legacy R4 settlement pivot** — first sheet has columns:
       ``state, lga, ward, settlement, total_treated``.
    """
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".csv")):
        raise HTTPException(400, "File must be .xlsx or .csv")
    raw = await file.read()

    # CSV path: treat as a single-sheet "Raw Data" or "legacy settlement
    # pivot" baseline. The multi-sheet R5 Target + Raw Data combo is xlsx-
    # only because CSV can't carry two sheets in one file.
    csv_rows: list = []
    wb = None
    if fname.endswith(".csv"):
        import csv as _csv
        try:
            text_blob = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text_blob = raw.decode("latin-1")
        csv_rows = list(_csv.reader(io.StringIO(text_blob)))
    else:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(400, f"Cannot open workbook: {e}")

    def strip_q(v):
        if v is None: return None
        s = str(v).strip().strip("'").strip()
        return s if s else None

    def to_int(v):
        if v is None: return None
        try: return int(round(float(str(v))))
        except Exception: return None

    def to_float(v):
        if v is None: return None
        try: return float(str(v))
        except Exception: return None

    def find_sheet(name_lc: str) -> str | None:
        if wb is None:
            return None
        return next((s for s in wb.sheetnames if s.strip().lower() == name_lc), None)

    def find_sheet_where(pred) -> str | None:
        if wb is None:
            return None
        return next((s for s in wb.sheetnames
                     if pred(" ".join(str(s).split()).strip().lower())), None)

    # The "target (ward)" sheet is named per round — "R5 Target (Ward)",
    # "Round 3 Target (Ward)", etc. — so match the stable words, not a fixed
    # round token (R5/R6/Round N all work). CSV has no sheets → None → the
    # single-sheet branch below handles it.
    target_sheet = find_sheet_where(lambda n: "target" in n and "ward" in n)
    raw_sheet = find_sheet("raw data") or find_sheet_where(lambda n: "raw" in n and "data" in n)

    # ── Pre-read Raw Data (if present) so the R5 Target path can use its
    #    f/m ratios for proportional split. Keyed by (LGA_upper, WARD_upper).
    raw_ratios: dict[tuple[str, str], dict] = {}
    if raw_sheet:
        ws_raw = wb[raw_sheet]
        rrows = list(ws_raw.iter_rows(values_only=True))
        rheader = [str(c or "").strip().lower() for c in (rrows[0] if rrows else [])]
        def rcol(n: str) -> int | None:
            try: return rheader.index(n)
            except ValueError: return None
        ri_lga, ri_ward = rcol("lga"), rcol("wardname")
        ri_1_11_f  = rcol("t_1_11_months_female")
        ri_1_11_m  = rcol("t_1_11_months_male")
        ri_12_59_f = rcol("t_12_59_months_female")
        ri_12_59_m = rcol("t_12_59_months_male")
        if None not in (ri_lga, ri_ward, ri_1_11_f, ri_1_11_m, ri_12_59_f, ri_12_59_m):
            for row in rrows[1:]:
                if ri_lga >= len(row): continue
                lga_v = strip_q(row[ri_lga])
                ward_v = strip_q(row[ri_ward]) if ri_ward < len(row) else None
                if not lga_v or not ward_v: continue
                raw_ratios[(lga_v.upper(), ward_v.upper())] = {
                    "f_1_11":  to_int(row[ri_1_11_f])  or 0,
                    "m_1_11":  to_int(row[ri_1_11_m])  or 0,
                    "f_12_59": to_int(row[ri_12_59_f]) or 0,
                    "m_12_59": to_int(row[ri_12_59_m]) or 0,
                }

    records: list = []  # (state, lga, ward, settlement, total, t1_11_f, t1_11_m, t12_59_f, t12_59_m)

    def split_band(total_band: int, f_raw: int, m_raw: int) -> tuple[int, int]:
        """Split a band's target into (female, male), preserving the sum.

        Uses Raw Data ratios when available, otherwise 50/50.
        """
        if total_band <= 0:
            return 0, 0
        denom = f_raw + m_raw
        if denom <= 0:
            f = total_band // 2
            return f, total_band - f
        f = int(round(total_band * f_raw / denom))
        if f > total_band: f = total_band
        return f, total_band - f

    if target_sheet:
        # ── Target (Ward) — authoritative totals ────────────────────────────
        ws = wb[target_sheet]
        rows = list(ws.iter_rows(values_only=True))
        # The target sheet uses multi-line headers; normalise to one line.
        header = [" ".join(str(c or "").split()).strip().lower() for c in (rows[0] if rows else [])]

        def col(name_lc: str) -> int | None:
            try: return header.index(name_lc)
            except ValueError: return None

        def find_col(pred) -> int | None:
            return next((idx for idx, h in enumerate(header) if pred(h)), None)

        # Fully round-agnostic. Columns are "<round> Target 1-11 Months",
        # "<round> Target 12-59 Months" and a grand-total "<round> Target" — the
        # round token ("R5", "Round 3", …) varies, so match the stable parts:
        # the band columns contain "N-M months"; the grand total is the column
        # that ends in "target" (the band columns end in "months").
        i_lga   = col("lga")
        i_ward  = col("ward")
        i_1_11  = find_col(lambda h: "1-11 months" in h)
        i_12_59 = find_col(lambda h: "12-59 months" in h)
        i_total = find_col(lambda h: h.endswith("target"))
        if None in (i_lga, i_ward, i_1_11, i_12_59, i_total):
            raise HTTPException(
                400,
                "The target sheet ('" + str(target_sheet) + "') is missing expected columns — need "
                "LGA, Ward, a '… 1-11 months' target, a '… 12-59 months' target, and a grand-total "
                "'… Target' column. Found: " + str(header),
            )

        for row in rows[1:]:
            if i_lga >= len(row): continue
            lga = strip_q(row[i_lga])
            ward = strip_q(row[i_ward]) if i_ward < len(row) else None
            if not lga or not ward: continue
            t_1_11_total  = to_int(row[i_1_11])  or 0
            t_12_59_total = to_int(row[i_12_59]) or 0
            grand_total   = to_int(row[i_total]) or (t_1_11_total + t_12_59_total)
            # Reconcile so band sums match grand total (rounding tolerance).
            if t_1_11_total + t_12_59_total != grand_total and (t_1_11_total + t_12_59_total) > 0:
                t_12_59_total = grand_total - t_1_11_total
                if t_12_59_total < 0:
                    t_1_11_total = grand_total
                    t_12_59_total = 0
            r = raw_ratios.get((lga.upper(), ward.upper()), {})
            f_1_11,  m_1_11  = split_band(t_1_11_total,  r.get("f_1_11", 0),  r.get("m_1_11", 0))
            f_12_59, m_12_59 = split_band(t_12_59_total, r.get("f_12_59", 0), r.get("m_12_59", 0))
            records.append(("Sokoto", lga.title(), ward.title(), None, grand_total,
                            f_1_11, m_1_11, f_12_59, m_12_59))
        chosen_format = "r5_target_ward"
    elif raw_sheet:
        # ── Raw Data only — direct M/F sum ──────────────────────────────────
        ws = wb[raw_sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c or "").strip().lower() for c in (rows[0] if rows else [])]

        def col(name: str) -> int | None:
            try: return header.index(name)
            except ValueError: return None

        i_lga, i_ward = col("lga"), col("wardname")
        i_1_11_f  = col("t_1_11_months_female")
        i_1_11_m  = col("t_1_11_months_male")
        i_12_59_f = col("t_12_59_months_female")
        i_12_59_m = col("t_12_59_months_male")

        if i_lga is None or i_ward is None or None in (i_1_11_f, i_1_11_m, i_12_59_f, i_12_59_m):
            raise HTTPException(400, f"'Raw Data' sheet is missing expected columns. Found: {header}")

        for row in rows[1:]:
            if i_lga >= len(row): continue
            lga = strip_q(row[i_lga])
            ward = strip_q(row[i_ward]) if i_ward < len(row) else None
            if not lga: continue
            t_1_11_f  = to_int(row[i_1_11_f])  or 0
            t_1_11_m  = to_int(row[i_1_11_m])  or 0
            t_12_59_f = to_int(row[i_12_59_f]) or 0
            t_12_59_m = to_int(row[i_12_59_m]) or 0
            total = t_1_11_f + t_1_11_m + t_12_59_f + t_12_59_m
            records.append(("Sokoto", lga.title(), ward.title() if ward else None,
                            None, total, t_1_11_f, t_1_11_m, t_12_59_f, t_12_59_m))
        chosen_format = "raw_data_only"
    else:
        # ── Legacy format: settlement pivot ────────────────────────────────
        # Single-sheet, also the path CSV uploads land in. Columns are now
        # resolved by header name (flexible matching) instead of positional
        # access, so a workbook with slightly different column ordering or
        # naming (e.g. "State Name" vs "state", "LGA Name" vs "lga") still
        # uploads cleanly.
        if wb is not None:
            ws = wb[wb.sheetnames[0]]
            all_rows = list(ws.iter_rows(values_only=True))
        else:
            all_rows = [tuple(r) for r in csv_rows] if csv_rows else []
        header_row   = list(all_rows[0]) if all_rows else []
        body_rows    = all_rows[1:] if len(all_rows) > 1 else []
        col_idx_base = _map_columns(header_row, {
            "state":      ("state_name", "state"),
            "lga":        ("lga_name", "lga"),
            "ward":       ("ward_name", "ward", "wardname"),
            "settlement": ("settlement_name", "settlement", "settlemen"),
            "total":      ("total_treated", "total", "target", "r5 target", "r4 target"),
        })

        def bcell(row, field):
            idx = col_idx_base.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row in body_rows:
            if not row:
                continue
            state = strip_q(bcell(row, "state"))
            lga   = strip_q(bcell(row, "lga"))
            ward  = strip_q(bcell(row, "ward"))
            sett  = strip_q(bcell(row, "settlement"))
            total = to_int(bcell(row, "total"))
            if lga and total is not None:
                records.append((state, lga, ward, sett, total, None, None, None, None))
        chosen_format = "legacy_settlement_pivot"

    if wb is not None:
        wb.close()

    conn = _get_sync_conn()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM mda_baseline WHERE project_id = %s", (pid,))
        if records:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO mda_baseline
                   (project_id, state, lga, ward, settlement, total_treated,
                    target_1_11_f, target_1_11_m, target_12_59_f, target_12_59_m)
                   VALUES %s""",
                [(pid, s, l, w, se, t, t1f, t1m, t2f, t2m) for s, l, w, se, t, t1f, t1m, t2f, t2m in records],
                page_size=500,
            )
        conn.commit()
    except Exception as e:
        conn.rollback(); raise HTTPException(500, f"DB error: {e}")
    finally:
        conn.close()

    return {
        "rows_inserted": len(records),
        "format": chosen_format,
    }
